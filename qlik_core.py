"""qlik_core - UI-independent Qlik Cloud logic.

Contains the engine/REST helpers, the QlikExporter (metadata extraction),
the cross-app consistency analysis, and the Excel workbook writer. No GUI
imports - shared by the Qt app and the headless CLI.
"""
import json
import csv
import os
import re
import datetime
import urllib.request
import urllib.parse
import urllib.error
import html
import websocket  # pip install websocket-client


def safe(s):
    return re.sub(r"[^A-Za-z0-9._-]+", "_", s).strip("_")


def normalize_host(tenant):
    return tenant.replace("https://", "").replace("wss://", "").strip("/")


def list_apps(tenant, api_key):
    """List all apps on the tenant via the Qlik Cloud REST items endpoint.
    Returns a sorted list of dicts {name, guid, space_id}. Handles pagination."""
    host = normalize_host(tenant)
    url = f"https://{host}/api/v1/items?resourceType=app&limit=100"
    apps = []
    while url:
        req = urllib.request.Request(url, headers={"Authorization": f"Bearer {api_key}"})
        with urllib.request.urlopen(req, timeout=30) as resp:
            data = json.loads(resp.read().decode("utf-8"))
        for it in data.get("data", []):
            guid = it.get("resourceId") or it.get("id")
            if not guid:
                continue
            apps.append({
                "name": it.get("name", "(unnamed)"),
                "guid": guid,
                "space_id": it.get("spaceId") or "",
            })
        nxt = (data.get("links", {}) or {}).get("next") or {}
        url = nxt.get("href")
    apps.sort(key=lambda a: (a["name"] or "").lower())
    return apps


def list_spaces(tenant, api_key):
    """List spaces on the tenant. Returns a dict of space_id -> space_name."""
    host = normalize_host(tenant)
    url = f"https://{host}/api/v1/spaces?limit=100"
    spaces = {}
    while url:
        req = urllib.request.Request(url, headers={"Authorization": f"Bearer {api_key}"})
        with urllib.request.urlopen(req, timeout=30) as resp:
            data = json.loads(resp.read().decode("utf-8"))
        for it in data.get("data", []):
            sid = it.get("id")
            if sid:
                spaces[sid] = it.get("name", sid)
        nxt = (data.get("links", {}) or {}).get("next") or {}
        url = nxt.get("href")
    return spaces


def list_data_files(tenant, api_key):
    """Best-effort: map data-file basename (lowercased) -> last modified date.
    Depends on tenant/connection access; the caller must handle failures."""
    host = normalize_host(tenant)
    url = f"https://{host}/api/v1/data-files?limit=100"
    out = {}
    while url:
        req = urllib.request.Request(url, headers={"Authorization": f"Bearer {api_key}"})
        with urllib.request.urlopen(req, timeout=30) as resp:
            data = json.loads(resp.read().decode("utf-8"))
        for it in data.get("data", []):
            base = os.path.basename(it.get("name") or "").lower()
            if base:
                out[base] = it.get("modifiedDate") or it.get("createdDate") or ""
        nxt = (data.get("links", {}) or {}).get("next") or {}
        url = nxt.get("href")
    return out


# ============================================================
#  Core export logic (UI-independent)
# ============================================================
class QlikExporter:
    def __init__(self, tenant, api_key, app_id, output_dir, log):
        self.tenant = normalize_host(tenant)
        self.api_key = api_key
        self.app_id = app_id
        self.output_dir = output_dir
        self.log = log
        self.id = 0
        self.ws = None

    # --- engine plumbing ---
    def connect(self):
        self.log(f"Connecting to {self.tenant} ...")
        self.ws = websocket.create_connection(
            f"wss://{self.tenant}/app/{self.app_id}",
            header=[f"Authorization: Bearer {self.api_key}"],
        )
        self.ws.recv()  # consume OnConnected
        self.log("Connected.")

    def call(self, handle, method, params):
        self.id += 1
        self.ws.send(json.dumps({
            "jsonrpc": "2.0", "id": self.id,
            "handle": handle, "method": method, "params": params,
        }))
        while True:
            msg = json.loads(self.ws.recv())
            if msg.get("id") == self.id:
                if "error" in msg:
                    raise RuntimeError(msg["error"])
                return msg["result"]

    def close(self):
        if self.ws:
            try:
                self.ws.close()
            except Exception:
                pass

    def _write_csv(self, path, rows):
        if not rows:
            self.log(f"  (nothing to write for {os.path.basename(path)})")
            return
        with open(path, "w", newline="", encoding="utf-8-sig") as f:
            w = csv.DictWriter(f, fieldnames=list(rows[0].keys()))
            w.writeheader()
            w.writerows(rows)

    # --- 1. master measures ---
    def fetch_measures(self, app_h):
        list_def = {"qInfo": {"qType": "MeasureList"},
                    "qMeasureListDef": {"qType": "measure"}}
        obj_h = self.call(app_h, "CreateSessionObject", [list_def])["qReturn"]["qHandle"]
        items = self.call(obj_h, "GetLayout", [])["qLayout"]["qMeasureList"]["qItems"]
        rows = []
        for it in items:
            mid = it["qInfo"]["qId"]
            mh = self.call(app_h, "GetMeasure", [mid])["qReturn"]["qHandle"]
            p = self.call(mh, "GetProperties", [])["qProp"]
            meta = p.get("qMetaDef", {})
            meas = p.get("qMeasure", {})
            fmt = meas.get("qNumFormat", {})
            rows.append({
                "id": mid,
                "name": meta.get("title", ""),
                "label": meas.get("qLabel", ""),
                "label_expression": meas.get("qLabelExpression", ""),
                "expression": meas.get("qDef", ""),
                "description": meta.get("description", ""),
                "tags": ", ".join(meta.get("tags", [])),
                "format_type": fmt.get("qType", ""),
                "format_pattern": fmt.get("qFmt", ""),
                "coloring": json.dumps(meas.get("coloring", {})) if meas.get("coloring") else "",
                "raw_json": json.dumps(p, ensure_ascii=False),
            })
        return rows

    def export_measures(self, app_h, prefix):
        rows = self.fetch_measures(app_h)
        path = os.path.join(self.output_dir, f"master_measures_{prefix}.csv")
        self._write_csv(path, rows)
        self.log(f"Exported {len(rows)} master measures -> {os.path.basename(path)}")

    # --- 2. master dimensions ---
    def fetch_dimensions(self, app_h):
        list_def = {"qInfo": {"qType": "DimensionList"},
                    "qDimensionListDef": {"qType": "dimension"}}
        obj_h = self.call(app_h, "CreateSessionObject", [list_def])["qReturn"]["qHandle"]
        items = self.call(obj_h, "GetLayout", [])["qLayout"]["qDimensionList"]["qItems"]
        rows = []
        for it in items:
            did = it["qInfo"]["qId"]
            dh = self.call(app_h, "GetDimension", [did])["qReturn"]["qHandle"]
            p = self.call(dh, "GetProperties", [])["qProp"]
            meta = p.get("qMetaDef", {})
            dim = p.get("qDim", {})
            rows.append({
                "id": did,
                "name": meta.get("title", ""),
                "grouping": dim.get("qGrouping", ""),          # N = single, H = drill-down
                "fields": "; ".join(dim.get("qFieldDefs", [])),
                "field_labels": "; ".join(dim.get("qFieldLabels", [])),
                "label_expression": dim.get("qLabelExpression", ""),
                "description": meta.get("description", ""),
                "tags": ", ".join(meta.get("tags", [])),
                "raw_json": json.dumps(p, ensure_ascii=False),
            })
        return rows

    def export_dimensions(self, app_h, prefix):
        rows = self.fetch_dimensions(app_h)
        path = os.path.join(self.output_dir, f"master_dimensions_{prefix}.csv")
        self._write_csv(path, rows)
        self.log(f"Exported {len(rows)} master dimensions -> {os.path.basename(path)}")

    # --- 3. variables ---
    def fetch_variables(self, app_h):
        list_def = {"qInfo": {"qType": "VariableList"},
                    "qVariableListDef": {"qType": "variable",
                                         "qShowReserved": True,
                                         "qShowConfig": True,
                                         "qData": {"tags": "/tags"}}}
        obj_h = self.call(app_h, "CreateSessionObject", [list_def])["qReturn"]["qHandle"]
        items = self.call(obj_h, "GetLayout", [])["qLayout"]["qVariableList"]["qItems"]
        rows = []
        for it in items:
            tags = (it.get("qData", {}) or {}).get("tags", []) or []
            rows.append({
                "id": it.get("qInfo", {}).get("qId", ""),
                "name": it.get("qName", ""),
                "definition": it.get("qDefinition", ""),
                "description": it.get("qDescription", ""),
                "is_script_created": it.get("qIsScriptCreated", False),
                "is_reserved": it.get("qIsReserved", False),
                "tags": ", ".join(tags),
                "raw_json": json.dumps(it, ensure_ascii=False),
            })
        return rows

    def export_variables(self, app_h, prefix):
        rows = self.fetch_variables(app_h)
        path = os.path.join(self.output_dir, f"variables_{prefix}.csv")
        self._write_csv(path, rows)
        self.log(f"Exported {len(rows)} variables -> {os.path.basename(path)}")

    # --- 4. load script ---
    def fetch_script(self, app_h):
        return self.call(app_h, "GetScript", [])["qScript"]

    def export_script(self, app_h, prefix):
        script = self.call(app_h, "GetScript", [])["qScript"]
        path = os.path.join(self.output_dir, f"load_script_{prefix}.qvs")
        with open(path, "w", encoding="utf-8-sig") as f:
            f.write(script)
        self.log(f"Exported load script ({len(script.splitlines())} lines) -> {os.path.basename(path)}")

    # --- 5. visuals ---
    @staticmethod
    def _walk(prop, sheet_title, rows):
        info = prop.get("qInfo", {})
        viz = prop.get("visualization") or info.get("qType", "")
        dims, meas = [], []
        hc = prop.get("qHyperCubeDef") or {}
        for d in hc.get("qDimensions", []):
            lib = d.get("qLibraryId", "")
            flds = d.get("qDef", {}).get("qFieldDefs", [])
            dims.append(lib or "; ".join(flds))
        for m in hc.get("qMeasures", []):
            lib = m.get("qLibraryId", "")
            expr = m.get("qDef", {}).get("qDef", "")
            meas.append(lib or expr)
        lo = prop.get("qListObjectDef") or {}
        if lo:
            lib = lo.get("qLibraryId", "")
            flds = lo.get("qDef", {}).get("qFieldDefs", [])
            dims.append(lib or "; ".join(flds))
        title = prop.get("title", "")
        if isinstance(title, dict):
            title = title.get("qStringExpression", {}).get("qExpr", "")
        if viz:
            rows.append({
                "sheet": sheet_title,
                "object_id": info.get("qId", ""),
                "type": viz,
                "title": title,
                "dimensions": " | ".join(d for d in dims if d),
                "measures": " | ".join(m for m in meas if m),
                "dim_count": len([d for d in dims if d]),
                "measure_count": len([m for m in meas if m]),
                "raw_json": json.dumps(prop, ensure_ascii=False),
            })

    def export_visuals(self, app_h, prefix):
        sheet_def = {
            "qInfo": {"qType": "SheetList"},
            "qAppObjectListDef": {
                "qType": "sheet",
                "qData": {"title": "/qMetaDef/title", "cells": "/cells"},
            },
        }
        sh = self.call(app_h, "CreateSessionObject", [sheet_def])["qReturn"]["qHandle"]
        sheets = self.call(sh, "GetLayout", [])["qLayout"]["qAppObjectList"]["qItems"]
        rows = []
        for s in sheets:
            sheet_title = s.get("qData", {}).get("title", s["qInfo"]["qId"])
            for cell in s.get("qData", {}).get("cells", []):
                obj_id = cell.get("name")
                if not obj_id:
                    continue
                try:
                    oh = self.call(app_h, "GetObject", [obj_id])["qReturn"]["qHandle"]
                    tree = self.call(oh, "GetFullPropertyTree", [])["qPropEntry"]
                except RuntimeError:
                    continue
                stack = [tree]
                while stack:
                    entry = stack.pop()
                    self._walk(entry.get("qProperty", {}), sheet_title, rows)
                    stack.extend(entry.get("qChildren", []) or [])
        path = os.path.join(self.output_dir, f"visuals_{prefix}.csv")
        self._write_csv(path, rows)
        self.log(f"Exported {len(rows)} visual objects across {len(sheets)} sheets -> {os.path.basename(path)}")

    # --- structured objects (for usage analysis) ---
    @staticmethod
    def _walk_struct(prop, sheet_title, rows):
        info = prop.get("qInfo", {})
        viz = prop.get("visualization") or info.get("qType", "")
        measure_libs, dim_libs, exprs = [], [], []
        hc = prop.get("qHyperCubeDef") or {}
        for d in hc.get("qDimensions", []):
            lib = d.get("qLibraryId", "")
            if lib:
                dim_libs.append(lib)
            else:
                exprs.extend(d.get("qDef", {}).get("qFieldDefs", []) or [])
        for m in hc.get("qMeasures", []):
            lib = m.get("qLibraryId", "")
            if lib:
                measure_libs.append(lib)
            else:
                expr = m.get("qDef", {}).get("qDef", "")
                if expr:
                    exprs.append(expr)
        lo = prop.get("qListObjectDef") or {}
        if lo:
            lib = lo.get("qLibraryId", "")
            if lib:
                dim_libs.append(lib)
            else:
                exprs.extend(lo.get("qDef", {}).get("qFieldDefs", []) or [])
        title = prop.get("title", "")
        if isinstance(title, dict):
            te = title.get("qStringExpression", {}).get("qExpr", "")
            if te:
                exprs.append(te)
            title = te
        if viz:
            rows.append({"sheet": sheet_title, "id": info.get("qId", ""), "type": viz,
                         "title": title if isinstance(title, str) else "",
                         "measure_libs": measure_libs, "dim_libs": dim_libs,
                         "expressions": [e for e in exprs if e]})

    def fetch_objects(self, app_h):
        sheet_def = {"qInfo": {"qType": "SheetList"},
                     "qAppObjectListDef": {"qType": "sheet",
                                           "qData": {"title": "/qMetaDef/title", "cells": "/cells"}}}
        sh = self.call(app_h, "CreateSessionObject", [sheet_def])["qReturn"]["qHandle"]
        sheets = self.call(sh, "GetLayout", [])["qLayout"]["qAppObjectList"]["qItems"]
        rows = []
        for s in sheets:
            sheet_title = s.get("qData", {}).get("title", s["qInfo"]["qId"])
            for cell in s.get("qData", {}).get("cells", []):
                obj_id = cell.get("name")
                if not obj_id:
                    continue
                try:
                    oh = self.call(app_h, "GetObject", [obj_id])["qReturn"]["qHandle"]
                    tree = self.call(oh, "GetFullPropertyTree", [])["qPropEntry"]
                except RuntimeError:
                    continue
                stack = [tree]
                while stack:
                    entry = stack.pop()
                    self._walk_struct(entry.get("qProperty", {}), sheet_title, rows)
                    stack.extend(entry.get("qChildren", []) or [])
        return rows

    def fetch_model_fields(self, app_h):
        list_def = {"qInfo": {"qType": "FieldList"},
                    "qFieldListDef": {"qShowSystem": True, "qShowHidden": True,
                                      "qShowSemantic": False, "qShowSrcTables": True,
                                      "qShowImplicit": False, "qShowDerivedFields": False}}
        obj_h = self.call(app_h, "CreateSessionObject", [list_def])["qReturn"]["qHandle"]
        items = self.call(obj_h, "GetLayout", [])["qLayout"]["qFieldList"]["qItems"]
        rows = []
        for it in items:
            tags = it.get("qTags", []) or []
            src = it.get("qSrcTables", []) or []
            rows.append({
                "name": it.get("qName", ""),
                "src_tables": src,
                "is_system": bool(it.get("qIsSystem", False)),
                "is_hidden": bool(it.get("qIsHidden", False)),
                "is_key": ("$key" in tags) or (len(src) > 1) or (it.get("qName", "").startswith("$Syn")),
                "tags": ", ".join(tags),
            })
        return rows

    def fetch_lineage(self, app_h):
        res = self.call(app_h, "GetLineage", [])
        return res.get("qLineage", []) or []

    # --- apply (write) master measures / dimensions ---
    def apply_master(self, app_h, kind, rows, mode, dry_run):
        """Create/update/delete master measures or dimensions from CSV rows.
        kind: 'measure' or 'dimension'. mode: 'upsert'/'create'/'update'/'delete'.
        Matches existing items by name (title). Returns a counts dict."""
        is_meas = (kind == "measure")
        list_type = "MeasureList" if is_meas else "DimensionList"
        list_key = "qMeasureListDef" if is_meas else "qDimensionListDef"
        layout_key = "qMeasureList" if is_meas else "qDimensionList"
        create_fn = "CreateMeasure" if is_meas else "CreateDimension"
        get_fn = "GetMeasure" if is_meas else "GetDimension"
        destroy_fn = "DestroyMeasure" if is_meas else "DestroyDimension"
        prop_key = "qMeasure" if is_meas else "qDim"

        list_def = {"qInfo": {"qType": list_type}, list_key: {"qType": kind}}
        obj_h = self.call(app_h, "CreateSessionObject", [list_def])["qReturn"]["qHandle"]
        items = self.call(obj_h, "GetLayout", [])["qLayout"][layout_key]["qItems"]
        existing = {it["qMeta"]["title"]: it["qInfo"]["qId"] for it in items}

        counts = {"created": 0, "updated": 0, "deleted": 0, "skipped": 0}
        for row in rows:
            name = (row.get("name") or "").strip()
            if not name:
                continue
            if mode == "delete":
                if name in existing:
                    self.log(f"DELETE {kind}: {name}")
                    if not dry_run:
                        self.call(app_h, destroy_fn, [existing[name]])
                    counts["deleted"] += 1
                else:
                    counts["skipped"] += 1
                continue
            body, meta = build_measure(row) if is_meas else build_dimension(row)
            if name in existing:
                if mode == "create":
                    counts["skipped"] += 1
                    continue
                self.log(f"UPDATE {kind}: {name}")
                if not dry_run:
                    h = self.call(app_h, get_fn, [existing[name]])["qReturn"]["qHandle"]
                    prop = self.call(h, "GetProperties", [])["qProp"]
                    prop.setdefault(prop_key, {}).update(body)
                    prop.setdefault("qMetaDef", {}).update(meta)
                    self.call(h, "SetProperties", [prop])
                counts["updated"] += 1
            else:
                if mode == "update":
                    counts["skipped"] += 1
                    continue
                self.log(f"CREATE {kind}: {name}")
                if not dry_run:
                    prop = {"qInfo": {"qType": kind}, prop_key: body, "qMetaDef": meta}
                    self.call(app_h, create_fn, [prop])
                counts["created"] += 1
        return counts

    def do_save(self, app_h):
        self.call(app_h, "DoSave", [])

    # --- orchestration ---
    def run(self, do_measures, do_dimensions, do_variables, do_script, do_visuals):
        os.makedirs(self.output_dir, exist_ok=True)
        self.connect()
        app_h = self.call(-1, "OpenDoc", [self.app_id])["qReturn"]["qHandle"]
        app_title = self.call(app_h, "GetAppLayout", [])["qLayout"].get("qTitle", self.app_id)
        stamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        prefix = f"{safe(app_title)}_{safe(self.app_id)}_{stamp}"
        self.log(f"App: {app_title}")
        if do_measures:
            self.export_measures(app_h, prefix)
        if do_dimensions:
            self.export_dimensions(app_h, prefix)
        if do_variables:
            self.export_variables(app_h, prefix)
        if do_script:
            self.export_script(app_h, prefix)
        if do_visuals:
            self.export_visuals(app_h, prefix)
        self.log("Done.")


# ============================================================
#  Apply (write) master items from CSV
# ============================================================
def read_csv_rows(path):
    with open(path, newline="", encoding="utf-8-sig") as f:
        return list(csv.DictReader(f))


def build_measure(row):
    meas = {"qDef": row.get("expression", "") or ""}
    if row.get("label"):
        meas["qLabel"] = row["label"]
    if row.get("label_expression"):
        meas["qLabelExpression"] = row["label_expression"]
    num = {}
    if row.get("format_type"):
        num["qType"] = row["format_type"]
    if row.get("format_pattern"):
        num["qFmt"] = row["format_pattern"]
    if num:
        meas["qNumFormat"] = num
    col = row.get("coloring")
    if col:
        try:
            meas["coloring"] = json.loads(col)
        except Exception:
            pass
    meta = {"title": row.get("name", "")}
    if row.get("description"):
        meta["description"] = row["description"]
    if row.get("tags"):
        meta["tags"] = [t.strip() for t in row["tags"].split(",") if t.strip()]
    return meas, meta


def build_dimension(row):
    fields = [s.strip() for s in (row.get("fields", "") or "").split(";") if s.strip()]
    dim = {"qGrouping": (row.get("grouping") or "N"), "qFieldDefs": fields}
    labels = [s.strip() for s in (row.get("field_labels", "") or "").split(";") if s.strip()]
    if labels:
        dim["qFieldLabels"] = labels
    if row.get("label_expression"):
        dim["qLabelExpression"] = row["label_expression"]
    meta = {"title": row.get("name", "")}
    if row.get("description"):
        meta["description"] = row["description"]
    if row.get("tags"):
        meta["tags"] = [t.strip() for t in row["tags"].split(",") if t.strip()]
    return dim, meta


# ============================================================
#  Cross-app consistency analysis (v1)
# ============================================================
def _tight(s):
    """Whitespace-insensitive, case-folded key for comparing definitions."""
    return re.sub(r"\s+", "", (s or "")).lower()


def expand_vars(expr, varmap, max_iter=6):
    """Best-effort substitution of $(var) references using a name->definition
    map. Parameterized ($(f(x))) or evaluated ($(=...)) refs are left as-is."""
    out = expr or ""
    for _ in range(max_iter):
        changed = False
        for name, defn in varmap.items():
            if not name:
                continue
            token = "$(" + name + ")"
            if token in out:
                out = out.replace(token, defn or "")
                changed = True
        if not changed:
            break
    return out


def _name_conflicts(items):
    """Items grouped by name where the calculation differs = consistency risk."""
    by_name = {}
    for it in items:
        nk = (it["name"] or "").strip().lower()
        by_name.setdefault(nk, []).append(it)
    out = []
    for group in by_name.values():
        variants = {}
        for it in group:
            v = variants.setdefault(it["calc_key"], {"expr": it["calc"], "apps": []})
            v["apps"].append(it["app"])
        if len(variants) > 1:
            out.append({
                "name": group[0]["name"],
                "variant_count": len(variants),
                "variants": [{"expr": v["expr"], "apps": sorted(set(v["apps"]))}
                             for v in variants.values()],
            })
    out.sort(key=lambda x: (-x["variant_count"], (x["name"] or "").lower()))
    return out


def _redundancy(items):
    """Items grouped by calculation that appear under more than one name."""
    by_calc = {}
    for it in items:
        if not it["calc_key"]:
            continue
        by_calc.setdefault(it["calc_key"], []).append(it)
    out = []
    for group in by_calc.values():
        names = sorted({(it["name"] or "") for it in group})
        apps = sorted({it["app"] for it in group})
        if len(names) > 1:
            out.append({"expr": group[0]["calc"], "names": names,
                        "apps": apps, "occurrences": len(group)})
    out.sort(key=lambda x: (-len(x["names"]), -x["occurrences"]))
    return out


def analyze_consistency(measures, dims):
    m_items = [{"name": m["name"],
                "calc": m.get("expr_expanded") or m["expression"],
                "calc_key": _tight(m.get("expr_expanded") or m["expression"]),
                "app": m["app"]} for m in measures]
    d_items = [{"name": d["name"],
                "calc": d.get("def_expanded") or d["definition"],
                "calc_key": _tight(d.get("def_expanded") or d["definition"]),
                "app": d["app"]} for d in dims]
    return {
        "measure_name_conflicts": _name_conflicts(m_items),
        "measure_redundancy": _redundancy(m_items),
        "dimension_name_conflicts": _name_conflicts(d_items),
        "dimension_redundancy": _redundancy(d_items),
    }


def write_consistency_report(results, measures, dims, out_dir, log):
    """Write the analysis to an Excel workbook (openpyxl). Falls back to a set
    of CSV files if openpyxl is not installed. Returns the main output path."""
    os.makedirs(out_dir, exist_ok=True)
    stamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    n_apps = len({m["app"] for m in measures} | {d["app"] for d in dims})
    base = f"consistency_report_{n_apps}apps_{stamp}"

    mnc, mr = results["measure_name_conflicts"], results["measure_redundancy"]
    dnc, dr = results["dimension_name_conflicts"], results["dimension_redundancy"]

    conflict_headers = ["Name", "Variant #", "Used in apps", "Apps", "Definition", "Unexpanded vars?"]

    def conflict_rows(conflicts):
        out = []
        for c in conflicts:
            for i, v in enumerate(c["variants"], 1):
                out.append([c["name"], i, len(v["apps"]), "; ".join(v["apps"]),
                            v["expr"], "Yes" if "$(" in (v["expr"] or "") else ""])
        return out

    redundancy_headers = ["Definition", "Distinct names", "Names", "Apps", "Occurrences"]

    def redundancy_rows(red):
        return [[r["expr"], len(r["names"]), "; ".join(r["names"]),
                 "; ".join(r["apps"]), r["occurrences"]] for r in red]

    m_inv_headers = ["App", "Space", "Measure name", "Definition (expanded)",
                     "Original definition", "Unexpanded vars?", "Tags"]
    m_inv_rows = [[m["app"], m.get("space", ""), m["name"],
                   m.get("expr_expanded") or m["expression"], m["expression"],
                   "Yes" if m.get("has_unexpanded") else "", m.get("tags", "")]
                  for m in sorted(measures, key=lambda x: ((x["name"] or "").lower(), x["app"]))]

    d_inv_headers = ["App", "Space", "Dimension name", "Grouping",
                     "Definition (expanded)", "Original definition", "Tags"]
    d_inv_rows = [[d["app"], d.get("space", ""), d["name"], d.get("grouping", ""),
                   d.get("def_expanded") or d["definition"], d["definition"], d.get("tags", "")]
                  for d in sorted(dims, key=lambda x: ((x["name"] or "").lower(), x["app"]))]

    summary_rows = [
        ["Apps analyzed", n_apps],
        ["Master measures scanned", len(measures)],
        ["Master dimensions scanned", len(dims)],
        ["Measure name conflicts (same name, different calc)", len(mnc)],
        ["Measure redundancy groups (same calc, different names)", len(mr)],
        ["Dimension name conflicts", len(dnc)],
        ["Dimension redundancy groups", len(dr)],
        ["Generated", stamp],
    ]

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except Exception:
        paths = []
        for suffix, headers, rows in [
            ("summary", ["Metric", "Value"], summary_rows),
            ("measure_name_conflicts", conflict_headers, conflict_rows(mnc)),
            ("measure_redundancy", redundancy_headers, redundancy_rows(mr)),
            ("dimension_name_conflicts", conflict_headers, conflict_rows(dnc)),
            ("dimension_redundancy", redundancy_headers, redundancy_rows(dr)),
            ("measures_all", m_inv_headers, m_inv_rows),
            ("dimensions_all", d_inv_headers, d_inv_rows),
        ]:
            p = os.path.join(out_dir, f"{base}_{suffix}.csv")
            with open(p, "w", newline="", encoding="utf-8-sig") as f:
                w = csv.writer(f)
                w.writerow(headers)
                w.writerows(rows)
            paths.append(p)
        log("openpyxl not installed - wrote CSV files instead of one workbook.")
        return paths[0]

    wb = Workbook()
    head_font = Font(bold=True, color="FFFFFF")
    head_fill = PatternFill("solid", fgColor="315C6D")
    wrap = Alignment(wrap_text=True, vertical="top")
    wide = {"Definition", "Definition (expanded)", "Original definition", "Apps", "Names"}
    first = {"done": False}

    def add_sheet(title, headers, rows, freeze=True):
        if not first["done"]:
            ws = wb.active
            first["done"] = True
        else:
            ws = wb.create_sheet()
        ws.title = title[:31]
        ws.append(headers)
        for c in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=c)
            cell.font = head_font
            cell.fill = head_fill
        for r in rows:
            rownum = ws.max_row + 1
            ws.append(r)
            # store any text that starts with "=" as a literal string, not a
            # formula - Qlik expressions begin with "=" and corrupt the file
            for ci, val in enumerate(r, 1):
                if isinstance(val, str) and val[:1] in ("=", "+", "-", "@"):
                    ws.cell(row=rownum, column=ci).data_type = "s"
        for i, h in enumerate(headers, 1):
            letter = get_column_letter(i)
            if h in wide:
                ws.column_dimensions[letter].width = 60
                for r in range(2, ws.max_row + 1):
                    ws.cell(row=r, column=i).alignment = wrap
            else:
                ws.column_dimensions[letter].width = max(12, min(40, len(h) + 4))
        if freeze and rows:
            ws.freeze_panes = "A2"
            ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"

    add_sheet("Summary", ["Metric", "Value"], summary_rows, freeze=False)
    add_sheet("Measure name conflicts", conflict_headers, conflict_rows(mnc))
    add_sheet("Measure redundancy", redundancy_headers, redundancy_rows(mr))
    add_sheet("Dimension name conflicts", conflict_headers, conflict_rows(dnc))
    add_sheet("Dimension redundancy", redundancy_headers, redundancy_rows(dr))
    add_sheet("Measures (all)", m_inv_headers, m_inv_rows)
    add_sheet("Dimensions (all)", d_inv_headers, d_inv_rows)

    out_path = os.path.join(out_dir, f"{base}.xlsx")
    wb.save(out_path)
    return out_path


# ============================================================
#  Usage / leanness analysis (what is NOT used) + dynamic-expr scan
# ============================================================
DOLLAR_RE = re.compile(r"\$\(([^)]*)\)")


def _referenced_names(text):
    names = set()
    if not text:
        return names
    for b in re.findall(r"\[([^\]]+)\]", text):
        names.add(b.strip().lower())
    for t in re.findall(r"[A-Za-z_][A-Za-z0-9_.]*", text):
        names.add(t.lower())
    return names


def _usage_corpus(measures, dimensions, variables, objects):
    texts = []
    for me in measures:
        texts.append(me.get("expression", ""))
        texts.append(me.get("label_expression", ""))
    for di in dimensions:
        texts.append(di.get("fields", ""))
        texts.append(di.get("label_expression", ""))
    for v in variables:
        texts.append(v.get("definition", ""))
    for o in objects:
        texts.extend(o.get("expressions", []) or [])
    return [t for t in texts if t]


def find_dynamic_expressions(measures, dimensions, variables, objects):
    out = []

    def scan(location, text):
        if not text:
            return
        for m in DOLLAR_RE.finditer(text):
            inner = m.group(1).strip()
            kind = "$(=...) active - HIDES USAGE" if inner.startswith("=") else "$(var) reference"
            out.append({"location": location, "type": kind, "expression": m.group(0)})

    for me in measures:
        scan(f"Master measure: {me.get('name', '')}", me.get("expression", ""))
        scan(f"Master measure label: {me.get('name', '')}", me.get("label_expression", ""))
    for di in dimensions:
        scan(f"Master dimension: {di.get('name', '')}", di.get("fields", ""))
        scan(f"Master dimension label: {di.get('name', '')}", di.get("label_expression", ""))
    for v in variables:
        scan(f"Variable: {v.get('name', '')}", v.get("definition", ""))
    for o in objects:
        loc = f"{o.get('type', 'object')} on sheet '{o.get('sheet', '')}'"
        for e in o.get("expressions", []) or []:
            scan(loc, e)
    return out


def analyze_master_usage(measures, dimensions, objects):
    use = {}
    for o in objects:
        for lib in (o.get("measure_libs", []) or []) + (o.get("dim_libs", []) or []):
            use.setdefault(lib, []).append(o)

    def rows_for(items, kind):
        res = []
        for it in items:
            objs = use.get(it.get("id", it.get("qId", "")), [])
            sheets = sorted({o.get("sheet", "") for o in objs})
            res.append({
                "kind": kind, "id": it.get("id", ""), "name": it.get("name", ""),
                "definition": it.get("expression", "") if kind == "measure" else it.get("fields", ""),
                "use_count": len(objs), "sheets": "; ".join(s for s in sheets if s),
                "used": bool(objs),
            })
        return res

    allrows = rows_for(measures, "measure") + rows_for(dimensions, "dimension")
    return {"all": allrows, "unused": [r for r in allrows if not r["used"]]}


def analyze_field_usage(model_fields, corpus):
    refs = set()
    for t in corpus:
        refs |= _referenced_names(t)
    rows = []
    for f in model_fields:
        nm = (f["name"] or "").lower()
        used = nm in refs
        if f["is_key"]:
            rec = "Key - keep"
        elif f["is_system"]:
            rec = "System - keep"
        elif f["is_hidden"]:
            rec = "Hidden - keep"
        else:
            rec = "Used" if used else "Unused candidate"
        rows.append({"name": f["name"], "tables": "; ".join(f["src_tables"]),
                     "is_key": f["is_key"], "is_system": f["is_system"], "is_hidden": f["is_hidden"],
                     "used": used, "recommendation": rec})
    return {"all": rows, "unused": [r for r in rows if r["recommendation"] == "Unused candidate"]}


def analyze_table_usage(model_fields, field_rows):
    used_by = {r["name"]: r["used"] for r in field_rows}
    tables = {}
    for f in model_fields:
        for tbl in (f["src_tables"] or ["(unknown)"]):
            t = tables.setdefault(tbl, {"fields": 0, "keys": 0, "data": 0, "used": 0})
            t["fields"] += 1
            if f["is_key"]:
                t["keys"] += 1
            elif f["is_system"] or f["is_hidden"]:
                pass  # don't count system/hidden as droppable data fields
            else:
                t["data"] += 1
                if used_by.get(f["name"]):
                    t["used"] += 1
    rows = []
    for name, t in sorted(tables.items()):
        flag = "No data fields used (review - may be a link table)" if (t["data"] > 0 and t["used"] == 0) else ""
        rows.append({"table": name, "fields": t["fields"], "key_fields": t["keys"],
                     "data_fields": t["data"], "used_data_fields": t["used"], "flag": flag})
    return rows


def analyze_variable_usage(variables, corpus):
    refs = set()
    for t in corpus:
        refs |= _referenced_names(t)
    rows = []
    for v in variables:
        if v.get("is_reserved"):
            continue
        nm = (v.get("name", "") or "").lower()
        used = nm in refs
        rows.append({"name": v.get("name", ""), "definition": v.get("definition", ""),
                     "is_script_created": v.get("is_script_created", False), "used": used})
    return {"all": rows, "unused": [r for r in rows if not r["used"]]}


def analyze_usage(measures, dimensions, variables, objects, model_fields):
    corpus = _usage_corpus(measures, dimensions, variables, objects)
    fields = analyze_field_usage(model_fields, corpus)
    return {
        "dynamic": find_dynamic_expressions(measures, dimensions, variables, objects),
        "master": analyze_master_usage(measures, dimensions, objects),
        "fields": fields,
        "tables": analyze_table_usage(model_fields, fields["all"]),
        "variables": analyze_variable_usage(variables, corpus),
    }


def write_usage_report(result, app_title, app_guid, out_dir, log):
    """Per-app usage/leanness workbook. Leads with a loud manual-review
    warning whenever the app uses dynamic $(...) expressions."""
    os.makedirs(out_dir, exist_ok=True)
    stamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    base = f"usage_report_{safe(app_title)}_{safe(app_guid)}_{stamp}"

    dyn = result["dynamic"]
    master = result["master"]
    fields = result["fields"]
    tables = result["tables"]
    variables = result["variables"]
    active = sum(1 for d in dyn if "active" in d["type"])

    warn = ("READ FIRST - results are CANDIDATES, not gospel. Usage is detected by "
            "text-matching expressions, so verify before deleting anything.")
    warn2 = (f"This app uses {len(dyn)} dynamic $(...) expressions ({active} active $(=...) ). "
             "Automated detection CANNOT see inside these - any field, master item or variable "
             "referenced only via a dynamic expression may be wrongly listed as unused. "
             "REVIEW the 'Dynamic expressions' sheet MANUALLY before removing anything.")

    summary_rows = [
        ["Model fields - unused candidates", len(fields["unused"])],
        ["Model fields - total (excl. keys/system are flagged)", len(fields["all"])],
        ["Master items - unused", len(master["unused"])],
        ["Master items - total", len(master["all"])],
        ["Variables - unused (front-end only)", len(variables["unused"])],
        ["Variables - total (excl. reserved)", len(variables["all"])],
        ["Tables flagged (no data fields used)", sum(1 for t in tables if t["flag"])],
        ["Dynamic $(...) expressions found", len(dyn)],
        ["  of which active $(=...)", active],
        ["Generated", stamp],
    ]

    dyn_headers = ["Location", "Type", "Expression"]
    dyn_rows = [[d["location"], d["type"], d["expression"]] for d in dyn]

    um_headers = ["Type", "Name", "Definition", "Id"]
    um_rows = [[r["kind"], r["name"], r["definition"], r["id"]] for r in master["unused"]]
    ma_headers = ["Type", "Name", "Used?", "Used in (# objects)", "Sheets", "Definition"]
    ma_rows = [[r["kind"], r["name"], "Yes" if r["used"] else "No", r["use_count"],
                r["sheets"], r["definition"]] for r in master["all"]]

    uf_headers = ["Field", "Table(s)"]
    uf_rows = [[r["name"], r["tables"]] for r in fields["unused"]]
    fa_headers = ["Field", "Table(s)", "Key?", "System?", "Hidden?", "Used?", "Recommendation"]
    fa_rows = [[r["name"], r["tables"], "Yes" if r["is_key"] else "", "Yes" if r["is_system"] else "",
                "Yes" if r["is_hidden"] else "", "Yes" if r["used"] else "No", r["recommendation"]]
               for r in fields["all"]]

    tb_headers = ["Table", "Fields", "Key fields", "Data fields", "Used data fields", "Flag"]
    tb_rows = [[t["table"], t["fields"], t["key_fields"], t["data_fields"], t["used_data_fields"], t["flag"]]
               for t in tables]

    uv_headers = ["Variable", "Definition", "Script-created?"]
    uv_rows = [[r["name"], r["definition"], "Yes" if r["is_script_created"] else ""]
               for r in variables["unused"]]

    sheets = [
        ("Dynamic expressions", dyn_headers, dyn_rows),
        ("Unused master items", um_headers, um_rows),
        ("Master item usage", ma_headers, ma_rows),
        ("Unused fields", uf_headers, uf_rows),
        ("Field usage", fa_headers, fa_rows),
        ("Table usage", tb_headers, tb_rows),
        ("Unused variables", uv_headers, uv_rows),
    ]

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except Exception:
        paths = []
        for suffix, headers, rows in [("summary", ["Metric", "Value"], summary_rows)] + \
                [(t.lower().replace(" ", "_"), h, r) for t, h, r in sheets]:
            pth = os.path.join(out_dir, f"{base}_{suffix}.csv")
            with open(pth, "w", newline="", encoding="utf-8-sig") as f:
                w = csv.writer(f)
                w.writerow(headers)
                w.writerows(rows)
            paths.append(pth)
        log("openpyxl not installed - wrote CSV files instead of one workbook.")
        return paths[0]

    wb = Workbook()
    head_font = Font(bold=True, color="FFFFFF")
    head_fill = PatternFill("solid", fgColor="315C6D")
    warn_font = Font(bold=True, color="B00020")
    wrap = Alignment(wrap_text=True, vertical="top")
    wide = {"Expression", "Definition", "Sheets", "Table(s)"}

    # Summary with the loud warning at the top
    ws = wb.active
    ws.title = "Summary"
    ws.append(["IMPORTANT - PLEASE READ"])
    ws["A1"].font = warn_font
    for line in (warn, warn2):
        ws.append([line])
        ws.cell(row=ws.max_row, column=1).font = warn_font
        ws.cell(row=ws.max_row, column=1).alignment = wrap
    ws.append([])
    ws.append(["Metric", "Value"])
    hdr = ws.max_row
    for c in (1, 2):
        ws.cell(row=hdr, column=c).font = head_font
        ws.cell(row=hdr, column=c).fill = head_fill
    for r in summary_rows:
        ws.append(r)
    ws.column_dimensions["A"].width = 70
    ws.column_dimensions["B"].width = 14

    def add_sheet(title, headers, rows):
        s = wb.create_sheet(title[:31])
        s.append(headers)
        for c in range(1, len(headers) + 1):
            s.cell(row=1, column=c).font = head_font
            s.cell(row=1, column=c).fill = head_fill
        for r in rows:
            rn = s.max_row + 1
            s.append(r)
            for ci, val in enumerate(r, 1):
                if isinstance(val, str) and val[:1] in ("=", "+", "-", "@"):
                    s.cell(row=rn, column=ci).data_type = "s"
        for i, h in enumerate(headers, 1):
            letter = get_column_letter(i)
            if h in wide:
                s.column_dimensions[letter].width = 55
                for r in range(2, s.max_row + 1):
                    s.cell(row=r, column=i).alignment = wrap
            else:
                s.column_dimensions[letter].width = max(12, min(36, len(h) + 4))
        if rows:
            s.freeze_panes = "A2"
            s.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{s.max_row}"

    for title, headers, rows in sheets:
        add_sheet(title, headers, rows)

    out_path = os.path.join(out_dir, f"{base}.xlsx")
    wb.save(out_path)
    return out_path


# ============================================================
#  Field lineage (backward: front end -> model table -> source)
# ============================================================
def _expr_uses_field(text, field_lower):
    return bool(field_lower) and field_lower in _referenced_names(text)


def _lineage_for_table(table, lineage):
    if not table:
        return []
    pat = re.compile(r"\b" + re.escape(table) + r"\b", re.I)
    hits = []
    for e in lineage:
        blob = (e.get("qDiscriminator") or "") + " " + (e.get("qStatement") or "")
        if pat.search(blob):
            hits.append(e)
    return hits


def trace_field_lineage(field_name, model_fields, lineage, measures, dimensions, variables, objects):
    fl = (field_name or "").strip().lower()
    fld = next((f for f in model_fields if (f.get("name") or "").lower() == fl), None)
    if not fld:
        return {"field": field_name, "found": False}

    mm_ref = [m for m in measures
              if _expr_uses_field(m.get("expression", ""), fl) or _expr_uses_field(m.get("label_expression", ""), fl)]
    md_ref = [d for d in dimensions
              if _expr_uses_field(d.get("fields", ""), fl) or _expr_uses_field(d.get("label_expression", ""), fl)]
    var_ref = [v for v in variables if _expr_uses_field(v.get("definition", ""), fl)]
    mm_by_id = {m.get("id"): m for m in mm_ref}
    md_by_id = {d.get("id"): d for d in md_ref}

    objs = []
    for o in objects:
        vias = []
        if any(_expr_uses_field(e, fl) for e in (o.get("expressions") or [])):
            vias.append("direct expression")
        for lib in o.get("measure_libs", []) or []:
            if lib in mm_by_id:
                vias.append(f"master measure '{mm_by_id[lib].get('name', '')}'")
        for lib in o.get("dim_libs", []) or []:
            if lib in md_by_id:
                vias.append(f"master dimension '{md_by_id[lib].get('name', '')}'")
        if vias:
            objs.append({"sheet": o.get("sheet", ""), "type": o.get("type", ""),
                         "via": "; ".join(sorted(set(vias)))})

    sources = []
    for tbl in (fld.get("src_tables") or []):
        hits = _lineage_for_table(tbl, lineage)
        sources.append({"table": tbl, "pinpointed": bool(hits),
                        "entries": [{"discriminator": e.get("qDiscriminator", ""),
                                     "statement": e.get("qStatement", "")} for e in hits]})

    return {
        "field": fld.get("name", field_name), "found": True,
        "is_key": fld.get("is_key", False), "is_system": fld.get("is_system", False),
        "tags": fld.get("tags", ""), "tables": fld.get("src_tables") or [],
        "consumed": {"measures": [m.get("name", "") for m in mm_ref],
                     "dimensions": [d.get("name", "") for d in md_ref],
                     "variables": [v.get("name", "") for v in var_ref],
                     "objects": objs},
        "sources": sources,
        "all_lineage": [{"discriminator": e.get("qDiscriminator", ""),
                         "statement": e.get("qStatement", "")} for e in lineage],
    }


def _parse_dt(s):
    if not s:
        return None
    try:
        return datetime.datetime.fromisoformat(str(s).replace("Z", "+00:00"))
    except Exception:
        return None


def extract_file_refs(text):
    if not text:
        return set()
    found = re.findall(r"([^\s\\/\[\]'\"]+\.(?:qvd|csv|txt|xlsx|xls|parquet))", text, re.I)
    return {os.path.basename(f).lower() for f in found}


def enrich_lineage_freshness(trace, app_last_reload, file_map):
    """Stamp app reload time and (best-effort) source-file modified dates.
    A source file newer than the app reload flags the app as possibly stale."""
    trace["app_last_reload"] = app_last_reload or ""
    arl = _parse_dt(app_last_reload)
    for s in trace.get("sources", []):
        blob = " ".join([s.get("table", "")] +
                        [(e.get("discriminator", "") + " " + e.get("statement", "")) for e in s.get("entries", [])])
        files = []
        for base in sorted(extract_file_refs(blob)):
            if base in file_map:
                mod = file_map[base]
                md = _parse_dt(mod)
                files.append({"name": base, "modified": mod,
                              "newer_than_reload": bool(md and arl and md > arl)})
        s["files"] = files
    return trace


def render_field_lineage_text(trace, app_title):
    if not trace.get("found"):
        return f"Field '{trace.get('field', '')}' was not found in the model of {app_title}."
    L = [f"FIELD LINEAGE   -   {trace['field']}     (app: {app_title})"]
    flags = []
    if trace.get("is_key"):
        flags.append("KEY field")
    if trace.get("is_system"):
        flags.append("SYSTEM field")
    if flags:
        L.append("flags: " + ", ".join(flags))
    if trace.get("app_last_reload"):
        L.append("app last reloaded: " + trace["app_last_reload"])
    c = trace["consumed"]
    L += ["", "CONSUMED BY (front end)"]
    if c["objects"]:
        for o in c["objects"]:
            L.append(f"  - {o['type'] or 'object'} on sheet '{o['sheet']}'   (via {o['via']})")
    else:
        L.append("  - no chart found using this field")
    if c["measures"]:
        L.append("  master measures: " + ", ".join(c["measures"]))
    if c["dimensions"]:
        L.append("  master dimensions: " + ", ".join(c["dimensions"]))
    if c["variables"]:
        L.append("  variables: " + ", ".join(c["variables"]))
    L += ["", "LIVES IN (model table)", "  " + (", ".join(trace["tables"]) or "(unknown)")]
    L += ["", "CAME FROM (source / load)"]
    unmatched = False
    for s in trace["sources"]:
        if s["pinpointed"]:
            L.append(f"  table '{s['table']}' loaded from:")
            for e in s["entries"]:
                L.append(f"     - {e['discriminator']}")
                if e["statement"]:
                    L.append("       " + " ".join(e["statement"].split())[:300])
        else:
            unmatched = True
            L.append(f"  table '{s['table']}': could not pinpoint a source statement - review the inventory below.")
        for f in s.get("files", []):
            tag = "   <-- NEWER than app reload: data may be STALE, reload the app" if f["newer_than_reload"] else ""
            L.append(f"       file '{f['name']}' modified: {f['modified'] or '(unknown)'}{tag}")
    if unmatched and trace["all_lineage"]:
        L += ["", "  App source inventory (all GetLineage entries):"]
        for e in trace["all_lineage"]:
            L.append(f"     - {e['discriminator']}")
    if trace.get("upstream"):
        L += ["", "UPSTREAM (apps that produce this field's source QVDs)"]
        for c in trace["upstream"]:
            ind = "  " * c["depth"]
            L.append(f"{ind}- QVD '{c['qvd']}' produced by app '{c['app']}' (last reload: {c['reload'] or 'unknown'})")
    elif trace.get("upstream_indexed"):
        L += ["", "UPSTREAM: no producing app found in the index for this field's QVDs",
              "  (the producer may not be among the loaded/indexed apps - load more and rebuild)."]
    L += ["", "NOTE: front-end usage is text-matched (verify dynamic $(=...) refs). Source mapping is at",
          "table/statement grain; exact source COLUMN and rename chain need load-script reading."]
    return "\n".join(L)


def write_field_lineage_html(trace, app_title, app_guid, out_dir):
    os.makedirs(out_dir, exist_ok=True)
    stamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    fname = f"lineage_{safe(trace.get('field', 'field'))}_{safe(app_title)}_{safe(app_guid)}_{stamp}.html"
    path = os.path.join(out_dir, fname)

    def esc(s):
        return html.escape(str(s or ""))

    def section(title):
        return f'<h2>{esc(title)}</h2>'

    parts = ['<!doctype html><meta charset="utf-8">',
             '<style>body{font-family:Segoe UI,Arial,sans-serif;color:#1F2A30;margin:24px;}'
             'h1{color:#315C6D;} h2{color:#315C6D;border-bottom:1px solid #D5DCDF;padding-bottom:4px;margin-top:24px;}'
             '.card{background:#fff;border:1px solid #D5DCDF;border-radius:8px;padding:14px 18px;}'
             'code{background:#F4F7F8;padding:1px 4px;border-radius:3px;} .muted{color:#5B6B72;}'
             '.flag{color:#B00020;font-weight:600;} ul{margin:6px 0;} pre{background:#F4F7F8;border:1px solid #D5DCDF;'
             'border-radius:6px;padding:8px;white-space:pre-wrap;font-size:12px;}</style>']
    if not trace.get("found"):
        parts.append(f"<h1>Field not found</h1><p>'{esc(trace.get('field',''))}' is not in the model of {esc(app_title)}.</p>")
        open(path, "w", encoding="utf-8").write("\n".join(parts))
        return path

    parts.append(f"<h1>Field lineage &middot; {esc(trace['field'])}</h1>")
    parts.append(f'<p class="muted">App: {esc(app_title)}</p>')
    if trace.get("app_last_reload"):
        parts.append(f'<p class="muted">App last reloaded: <b>{esc(trace["app_last_reload"])}</b></p>')
    fl = []
    if trace.get("is_key"):
        fl.append("KEY field")
    if trace.get("is_system"):
        fl.append("SYSTEM field")
    if fl:
        parts.append(f'<p class="flag">Flags: {esc(", ".join(fl))}</p>')

    parts.append('<div class="card">')
    parts.append(section("Consumed by (front end)"))
    c = trace["consumed"]
    if c["objects"]:
        parts.append("<ul>")
        for o in c["objects"]:
            parts.append(f"<li>{esc(o['type'] or 'object')} on sheet <b>{esc(o['sheet'])}</b> "
                         f"<span class='muted'>(via {esc(o['via'])})</span></li>")
        parts.append("</ul>")
    else:
        parts.append("<p class='muted'>No chart found using this field.</p>")
    for lbl, key in (("Master measures", "measures"), ("Master dimensions", "dimensions"), ("Variables", "variables")):
        if c[key]:
            parts.append(f"<p><b>{lbl}:</b> {esc(', '.join(c[key]))}</p>")
    parts.append("</div>")

    parts.append('<div class="card">')
    parts.append(section("Lives in (model table)"))
    parts.append(f"<p>{esc(', '.join(trace['tables']) or '(unknown)')}</p>")
    parts.append("</div>")

    parts.append('<div class="card">')
    parts.append(section("Came from (source / load)"))
    unmatched = False
    for s in trace["sources"]:
        if s["pinpointed"]:
            parts.append(f"<p>Table <b>{esc(s['table'])}</b> loaded from:</p><ul>")
            for e in s["entries"]:
                parts.append(f"<li><code>{esc(e['discriminator'])}</code>")
                if e["statement"]:
                    parts.append(f"<pre>{esc(e['statement'])}</pre>")
                parts.append("</li>")
            parts.append("</ul>")
        else:
            unmatched = True
            parts.append(f"<p class='flag'>Table {esc(s['table'])}: could not pinpoint a source statement - review the inventory.</p>")
        for f in s.get("files", []):
            if f["newer_than_reload"]:
                parts.append(f"<p class='flag'>Source file <code>{esc(f['name'])}</code> modified {esc(f['modified'])} "
                             "- NEWER than the app reload, so the app may be showing stale data (reload it).</p>")
            else:
                parts.append(f"<p class='muted'>Source file <code>{esc(f['name'])}</code> modified {esc(f['modified'] or '(unknown)')}.</p>")
    if unmatched and trace["all_lineage"]:
        parts.append("<p class='muted'>App source inventory (all GetLineage entries):</p><ul>")
        for e in trace["all_lineage"]:
            parts.append(f"<li><code>{esc(e['discriminator'])}</code></li>")
        parts.append("</ul>")
    parts.append("</div>")

    if trace.get("upstream"):
        parts.append('<div class="card">')
        parts.append(section("Upstream (apps that produce this field's source QVDs)"))
        parts.append("<ul>")
        for c in trace["upstream"]:
            pad = 18 * (c["depth"] - 1)
            parts.append(f"<li style='margin-left:{pad}px'>QVD <code>{esc(c['qvd'])}</code> produced by app "
                         f"<b>{esc(c['app'])}</b> <span class='muted'>(last reload: {esc(c['reload'] or 'unknown')})</span></li>")
        parts.append("</ul></div>")
    elif trace.get("upstream_indexed"):
        parts.append("<p class='muted'>Upstream: no producing app found in the index for this field's QVDs "
                     "(the producer may not be among the loaded/indexed apps).</p>")
    parts.append('<p class="muted" style="margin-top:18px;">Front-end usage is text-matched - verify dynamic '
                 '$(=...) references. Source mapping is at table/statement grain; exact source column and rename '
                 'chain need load-script reading (not done in this version).</p>')
    open(path, "w", encoding="utf-8").write("\n".join(parts))
    return path


# ============================================================
#  Cross-app lineage (which app STOREs which QVD -> walk upstream)
# ============================================================
def parse_store_reads(script):
    """From a load script, return (stores, reads): basenames of QVDs the app
    writes (STORE ... INTO ...qvd) and QVDs it reads (everything else)."""
    if not script:
        return set(), set()
    stores = set()
    for m in re.finditer(r"\bSTORE\b[^;]*?\bINTO\b\s*([^;()]+\.qvd)", script, re.I | re.S):
        base = os.path.basename(m.group(1).strip()).strip("[]' ").replace('"', "").lower()
        if base:
            stores.add(base)
    all_qvd = {b for b in extract_file_refs(script) if b.endswith(".qvd")}
    return stores, (all_qvd - stores)


def build_producer_map(index):
    """index: list of {guid, name, reload, stores:[...], reads:[...]}.
    Returns qvd_basename -> list of producing app entries."""
    pm = {}
    for app in index:
        for q in app.get("stores", []) or []:
            pm.setdefault(q, []).append(app)
    return pm


def trace_upstream_chain(start_qvds, producer_map, max_depth=6):
    chain = []
    seen = set()
    frontier = [(1, q) for q in sorted(start_qvds)]
    while frontier:
        depth, qvd = frontier.pop(0)
        if depth > max_depth:
            continue
        for p in producer_map.get(qvd, []):
            chain.append({"depth": depth, "qvd": qvd, "app": p.get("name", ""),
                          "guid": p.get("guid", ""), "reload": p.get("reload", "")})
            if p.get("guid") not in seen:
                seen.add(p.get("guid"))
                for inq in p.get("reads", []) or []:
                    frontier.append((depth + 1, inq))
    return chain


def attach_upstream(trace, producer_map):
    trace["upstream_indexed"] = producer_map is not None
    if not producer_map:
        trace["upstream"] = []
        return trace
    start = set()
    for s in trace.get("sources", []):
        for e in s.get("entries", []):
            for b in extract_file_refs((e.get("discriminator", "") + " " + e.get("statement", ""))):
                if b.endswith(".qvd"):
                    start.add(b)
    trace["upstream"] = trace_upstream_chain(start, producer_map)
    return trace


# ============================================================
#  Qlik native lineage (lineage-graphs REST) -> consumer provenance
# ============================================================
def fetch_native_lineage(tenant, api_key, app_guid, level="all", up=-1, collapse=True):
    """Call Qlik's lineage-graphs API for an app and return parsed
    {nodes: {qri: meta}, edges: [{source,target,relation}]}.
    Raises on HTTP error so the caller can fall back."""
    host = normalize_host(tenant)
    qri = "qri:app:sense://" + app_guid
    enc = urllib.parse.quote(qri, safe="")
    url = (f"https://{host}/api/v1/lineage-graphs/nodes/{enc}"
           f"?level={level}&up={up}&collapse={'true' if collapse else 'false'}")
    req = urllib.request.Request(url, headers={"Authorization": "Bearer " + api_key})
    with urllib.request.urlopen(req, timeout=60) as resp:
        data = json.loads(resp.read().decode("utf-8"))

    graphs = []
    if isinstance(data.get("graph"), dict):
        graphs = [data["graph"]]
    elif isinstance(data.get("graphs"), dict):
        graphs = data["graphs"].get("graphs", []) or []
    elif isinstance(data.get("graphs"), list):
        graphs = data["graphs"]

    nodes, edges = {}, []
    for g in graphs:
        raw = g.get("nodes")
        if isinstance(raw, str):
            try:
                raw = json.loads(raw)
            except Exception:
                raw = {}
        if isinstance(raw, dict):
            for nid, val in raw.items():
                meta = dict(val.get("metadata", {})) if isinstance(val, dict) else {}
                meta["label"] = (val.get("label", "") if isinstance(val, dict) else "") or meta.get("label", "")
                nodes[nid] = meta
        for e in g.get("edges", []) or []:
            edges.append({"source": e.get("source"), "target": e.get("target"),
                          "relation": e.get("relation", []) or []})
    return {"nodes": nodes, "edges": edges}


def _native_node_kind(qri, meta):
    t = (meta.get("type") or "").upper()
    fp = meta.get("filePath") or ""
    q = (qri or "").lower()
    if t == "FIELD":
        return "Field"
    if t == "TABLE":
        return "Table"
    if q.startswith("qri:db:") or "DATABASE" in t:
        return "Database"
    if fp:
        return "File (" + os.path.basename(fp) + ")"
    if q.startswith("qri:app:") or t == "DA_APP":
        return "Qlik app"
    if "DATASET" in t:
        return "Dataset"
    return (t.title().replace("_", " ") or "Source")


def native_provenance(graph, field_name, app_guid=None):
    """Walk Qlik's lineage graph upstream to a field's origins. Tries the
    field node first; if its path reaches no source resource, falls back to
    the report app's own upstream sources. Direction-agnostic."""
    nodes = graph.get("nodes", {})
    edges = graph.get("edges", [])
    preds, succs = {}, {}
    for e in edges:
        s, t = e.get("source"), e.get("target")
        if s and t:
            preds.setdefault(t, []).append(s)
            succs.setdefault(s, []).append(t)

    def walk(starts):
        def bfs(adj):
            seen, order = set(), []
            frontier = [(0, q) for q in starts]
            while frontier:
                depth, q = frontier.pop(0)
                if q in seen:
                    continue
                seen.add(q)
                m = nodes.get(q, {})
                order.append({"qri": q, "label": m.get("label", "") or q,
                              "kind": _native_node_kind(q, m), "depth": depth})
                for nxt in adj.get(q, []):
                    if nxt not in seen:
                        frontier.append((depth + 1, nxt))
            return order
        order = bfs(preds)
        if len(order) <= len(starts) and any(succs.get(s) for s in starts):
            order = bfs(succs)
        return order

    def has_sources(chain):
        return any(c["kind"].startswith(("Database", "Qlik app", "File", "Dataset")) for c in chain)

    fl = (field_name or "").strip().lower()
    field_starts = [q for q, m in nodes.items() if m.get("label", "").strip().lower() == fl]
    found_field = bool(field_starts)

    chain = walk(field_starts) if field_starts else []
    mode = "field"
    if not has_sources(chain):
        # fall back to the report app's own upstream sources
        app_starts = []
        if app_guid:
            target = ("qri:app:sense://" + app_guid).lower()
            app_starts = [q for q in nodes if str(q).lower() == target]
        if not app_starts:
            app_starts = [q for q in nodes if str(q).lower().startswith("qri:app:")]
        app_chain = walk(app_starts)
        if has_sources(app_chain):
            chain = app_chain
            mode = "app"

    return {"field": field_name, "found_field": found_field, "mode": mode, "chain": chain}


def _native_resource_chain(prov, app_title=""):
    keep = ("Database", "Qlik app", "File", "Dataset")
    at = (app_title or "").strip().lower()
    seen, res = set(), []
    for c in prov.get("chain", []):
        if c["kind"].startswith(keep) and c["label"] not in seen and c["label"].strip().lower() != at:
            seen.add(c["label"])
            res.append(c)
    res.sort(key=lambda x: -x["depth"])   # furthest source first
    return res


def render_native_provenance_text(prov, app_title):
    if not prov.get("chain"):
        return f"No Qlik lineage was returned for '{prov.get('field', '')}' in {app_title}."
    L = [f"DATA PROVENANCE   -   {prov.get('field', '')}     (report: {app_title})"]
    if not prov.get("found_field") or prov.get("mode") == "app":
        L.append("(could not pin this exact field's path - showing the report's overall data sources)")
    L += ["", "Where the values come from (furthest source first):"]
    res = _native_resource_chain(prov, app_title)
    if res:
        for c in res:
            L.append(f"  [{c['kind']}] {c['label']}")
        L.append(f"  [This report] {app_title}")
    else:
        L.append("  (no upstream sources resolved)")
    L += ["", "Source: Qlik's own data lineage (lineage-graphs API)."]
    return "\n".join(L)


def write_native_provenance_html(prov, app_title, app_guid, out_dir):
    os.makedirs(out_dir, exist_ok=True)
    stamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    fname = f"provenance_{safe(prov.get('field', 'field'))}_{safe(app_title)}_{safe(app_guid)}_{stamp}.html"
    path = os.path.join(out_dir, fname)

    def esc(s):
        return html.escape(str(s or ""))

    res = _native_resource_chain(prov, app_title)
    parts = ['<!doctype html><meta charset="utf-8">',
             '<style>body{font-family:Segoe UI,Arial,sans-serif;color:#1F2A30;margin:24px;}'
             'h1{color:#315C6D;} .muted{color:#5B6B72;} .card{background:#fff;border:1px solid #D5DCDF;'
             'border-radius:8px;padding:16px 20px;max-width:760px;}'
             '.step{padding:8px 12px;border:1px solid #D5DCDF;border-radius:8px;margin:6px 0;background:#F4F7F8;}'
             '.kind{display:inline-block;background:#315C6D;color:#fff;border-radius:10px;padding:1px 8px;'
             'font-size:11px;margin-right:8px;} .arrow{color:#5B6B72;text-align:center;margin:2px 0;}</style>']
    parts.append(f"<h1>Where does &ldquo;{esc(prov.get('field',''))}&rdquo; come from?</h1>")
    parts.append(f'<p class="muted">Report: {esc(app_title)}</p>')
    if not prov.get("found_field"):
        parts.append("<p class='muted'>(Showing the overall data sources for this report.)</p>")
    parts.append('<div class="card">')
    if res:
        for i, c in enumerate(res):
            tag = "ORIGIN" if i == 0 else c["kind"]
            parts.append(f'<div class="step"><span class="kind">{esc(c["kind"])}</span>{esc(c["label"])}</div>')
            parts.append('<div class="arrow">&darr;</div>')
        parts.append(f'<div class="step"><span class="kind">Report</span>{esc(app_title)}</div>')
    else:
        parts.append('<p class="muted">No upstream sources resolved.</p>')
    parts.append('</div>')
    parts.append('<p class="muted" style="margin-top:16px;">Source: Qlik&rsquo;s own data lineage '
                 '(lineage-graphs API).</p>')
    open(path, "w", encoding="utf-8").write("\n".join(parts))
    return path


def write_lineage_graph_html(prov, app_title, app_guid, out_dir):
    """Render the provenance as a layered SVG graph (origin -> ... -> report)
    in a self-contained, offline HTML file."""
    os.makedirs(out_dir, exist_ok=True)
    stamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    fname = f"lineage_graph_{safe(prov.get('field', 'field'))}_{safe(app_title)}_{safe(app_guid)}_{stamp}.html"
    path = os.path.join(out_dir, fname)

    def esc(s):
        return html.escape(str(s or ""))

    res = _native_resource_chain(prov, app_title)
    depths = sorted({n["depth"] for n in res}, reverse=True)
    rows = [[n for n in res if n["depth"] == d] for d in depths]
    rows.append([{"label": app_title, "kind": "Report"}])
    rows = [r for r in rows if r] or [[{"label": app_title, "kind": "Report"}]]

    BW, BH, HG, VG, MX, MY = 220, 54, 44, 88, 30, 30
    maxcols = max(len(r) for r in rows)
    width = MX * 2 + maxcols * BW + (maxcols - 1) * HG
    height = MY * 2 + len(rows) * BH + (len(rows) - 1) * VG

    def row_y(i):
        return MY + i * (BH + VG)

    def node_x(rowlen, j):
        row_w = rowlen * BW + (rowlen - 1) * HG
        startx = (width - row_w) / 2
        return startx + j * (BW + HG)

    palette = [("Database", "#7a5230"), ("File", "#5B6B72"), ("Qlik app", "#315C6D"),
               ("Dataset", "#3E7185"), ("Report", "#274A57")]

    def color(kind):
        for k, c in palette:
            if kind.startswith(k):
                return c
        return "#5B6B72"

    svg = [f'<svg width="{width}" height="{height}" xmlns="http://www.w3.org/2000/svg" '
           'font-family="Segoe UI, Arial, sans-serif">',
           '<defs><marker id="arrow" markerWidth="10" markerHeight="10" refX="8" refY="3" '
           'orient="auto" markerUnits="strokeWidth"><path d="M0,0 L8,3 L0,6 Z" fill="#9aa6ab"/></marker></defs>']
    # edges between consecutive layers (data flows downward)
    for i in range(len(rows) - 1):
        for j, _ in enumerate(rows[i]):
            ax = node_x(len(rows[i]), j) + BW / 2
            ay = row_y(i) + BH
            for k, _b in enumerate(rows[i + 1]):
                bx = node_x(len(rows[i + 1]), k) + BW / 2
                by = row_y(i + 1)
                svg.append(f'<line x1="{ax:.0f}" y1="{ay:.0f}" x2="{bx:.0f}" y2="{by:.0f}" '
                           'stroke="#9aa6ab" stroke-width="1.5" marker-end="url(#arrow)"/>')
    # boxes on top
    for i, row in enumerate(rows):
        for j, n in enumerate(row):
            x = node_x(len(row), j)
            y = row_y(i)
            lab = n["label"]
            lab = (lab[:28] + "…") if len(lab) > 28 else lab
            svg.append(f'<rect x="{x:.0f}" y="{y}" width="{BW}" height="{BH}" rx="10" fill="{color(n["kind"])}"/>')
            svg.append(f'<text x="{x + BW / 2:.0f}" y="{y + 20}" fill="#cfe0e6" font-size="10" '
                       f'text-anchor="middle">{esc(n["kind"])}</text>')
            svg.append(f'<text x="{x + BW / 2:.0f}" y="{y + 39}" fill="#ffffff" font-size="13" '
                       f'font-weight="600" text-anchor="middle">{esc(lab)}</text>')
    svg.append('</svg>')

    page = ['<!doctype html><meta charset="utf-8">',
            '<style>body{font-family:Segoe UI,Arial,sans-serif;color:#1F2A30;margin:24px;background:#EEF1F3;}'
            'h1{color:#315C6D;} .muted{color:#5B6B72;} .wrap{background:#fff;border:1px solid #D5DCDF;'
            'border-radius:10px;padding:16px;display:inline-block;}</style>',
            f"<h1>Where does &ldquo;{esc(prov.get('field', ''))}&rdquo; come from?</h1>",
            f'<p class="muted">Report: {esc(app_title)} &middot; flow reads top (origin) to bottom (report)</p>',
            '<div class="wrap">', "".join(svg), '</div>',
            '<p class="muted" style="margin-top:14px;">Source: Qlik&rsquo;s own data lineage (lineage-graphs API).</p>']
    open(path, "w", encoding="utf-8").write("\n".join(page))
    return path


# ============================================================
#  Field-specific load-script pipeline  (slim "where did THIS
#  field come from" - the path it took INTO this app)
#
#  Goal: instead of dumping every source the whole app touches,
#  parse the load script, find the LOAD that produces the field,
#  and follow it back to its ORIGIN:
#    - if the field was loaded FROM an external source (qvd/file/db)
#      -> show that external source (the real origin), or
#    - if it was loaded RESIDENT from another in-app table
#      -> walk the resident chain back until an external FROM is hit
#         (otherwise report the internal resident table).
#  Only the pipeline is shown - not every place the field is used.
# ============================================================
def _strip_script_comments(script):
    """Remove /* */ blocks, REM ...; and // line comments - but keep the
    '//' inside lib:// / http:// paths."""
    if not script:
        return ""
    s = re.sub(r"/\*.*?\*/", " ", script, flags=re.S)
    s = re.sub(r"(?im)\bREM\b[^;]*;", " ", s)
    out = []
    for line in s.split("\n"):
        # cut a line comment only when '//' is NOT preceded by ':' (lib://)
        m = re.search(r"(?<!:)//", line)
        out.append(line[:m.start()] if m else line)
    return "\n".join(out)


def _split_statements(script):
    """Split a load script into statements on ';', honouring [...] brackets
    (lib paths, INLINE data) and quotes so embedded ';' do not split."""
    stmts, buf = [], []
    depth = 0
    quote = None
    for ch in script:
        if quote:
            buf.append(ch)
            if ch == quote:
                quote = None
            continue
        if ch in ("'", '"'):
            quote = ch
            buf.append(ch)
            continue
        if ch == "[":
            depth += 1
            buf.append(ch)
            continue
        if ch == "]":
            depth = max(0, depth - 1)
            buf.append(ch)
            continue
        if ch == ";" and depth == 0:
            s = "".join(buf).strip()
            if s:
                stmts.append(s)
            buf = []
            continue
        buf.append(ch)
    tail = "".join(buf).strip()
    if tail:
        stmts.append(tail)
    return stmts


def _split_top_level(s, sep=","):
    """Split on sep at top level (paren / bracket / quote aware)."""
    parts, buf = [], []
    depth = 0
    quote = None
    for ch in s:
        if quote:
            buf.append(ch)
            if ch == quote:
                quote = None
            continue
        if ch in ("'", '"'):
            quote = ch
            buf.append(ch)
            continue
        if ch in "([":
            depth += 1
        elif ch in ")]":
            depth = max(0, depth - 1)
        if ch == sep and depth == 0:
            parts.append("".join(buf))
            buf = []
            continue
        buf.append(ch)
    parts.append("".join(buf))
    return parts


def _clean_ident(tok):
    """Strip brackets/quotes/whitespace from a field or table identifier."""
    t = (tok or "").strip()
    if len(t) >= 2 and t[0] in "[\"'" and t[-1] in "]\"'":
        t = t[1:-1]
    return t.strip()


def _field_in_out_name(piece):
    """Given one comma-separated LOAD field expression, return (out_name, in_name).
    out_name = the field this LOAD produces; in_name = the single source field it
    renames from, or None if the left side is a real expression (not traceable)."""
    # find a top-level ' AS '
    m = None
    for mm in re.finditer(r"(?i)\bas\b", piece):
        # ensure not inside parens: count parens before it
        before = piece[:mm.start()]
        if before.count("(") == before.count(")"):
            m = mm
    if m:
        lhs = piece[:m.start()]
        rhs = piece[m.end():]
        out = _clean_ident(_split_top_level(rhs)[0])
        lhs_str = lhs.strip()
        if re.fullmatch(r"[\[\"']?[\w. ]+[\]\"']?", lhs_str):
            return out, _clean_ident(lhs_str)
        return out, None
    bare = _clean_ident(piece)
    return bare, (bare or None)


_SRC_KW = re.compile(r"(?i)\b(FROM|RESIDENT|INLINE|AUTOGENERATE)\b")


def _sql_source_label(sql_text, conn):
    """Describe a SQL source as '<connection> · <table>' (best-effort)."""
    mfrom = re.search(r"(?is)\bFROM\b\s+([A-Za-z0-9_.\[\]\"`]+)", sql_text or "")
    tbl = _clean_ident(mfrom.group(1)) if mfrom else ""
    conn = (conn or "").strip()
    if conn and tbl:
        return f"{conn}  ·  {tbl}"
    if tbl:
        return "Database table " + tbl
    if conn:
        return "Database " + conn
    return "Database (SQL SELECT)"


def parse_load_tables(script):
    """Parse the load script into table-producing LOAD/SQL statements.
    Returns a list of dicts:
      {label, out_fields:[{out,in}], out_lc:set, wildcard:bool,
       kind: from|resident|inline|autogenerate|sql|unknown,
       source: display string, files:set(basenames), external:bool, raw}"""
    clean = _strip_script_comments(script or "")
    stmts = _split_statements(clean)
    tables = []
    cur_conn = ""   # most recent data connection (LIB/ODBC/OLEDB CONNECT TO ...)
    for idx, st in enumerate(stmts):
        cm = re.search(r"(?is)\bCONNECT\s+TO\s+('([^']+)'|\[([^\]]+)\]|\"([^\"]+)\"|([^\s;(]+))", st)
        if cm:
            cur_conn = next((g for g in cm.groups()[1:] if g), "") or cur_conn
        body = st
        # leading label  "Name:"  / "[Name]:" / "\"Name\":"
        label = ""
        lm = re.match(r"""(?s)^\s*(\[[^\]]+\]|"[^"]+"|'[^']+'|[A-Za-z_][\w]*)\s*:\s*(.*)$""", body)
        if lm and re.search(r"(?is)\bLOAD\b|\bSQL\b", lm.group(2) or ""):
            label = _clean_ident(lm.group(1))
            body = lm.group(2)
        # strip prefixes that sit before LOAD (best-effort)
        prefix = re.match(r"(?is)^\s*((?:NO)?CONCATENATE|(?:LEFT|RIGHT|INNER\s+)?JOIN|"
                          r"MAPPING|BUFFER|ADD|REPLACE|SEMANTIC|HIERARCHY|GENERIC)\b(\s*\([^)]*\))?\s*(.*)$",
                          body)
        join_target = ""
        if prefix:
            if prefix.group(2):
                join_target = _clean_ident(prefix.group(2).strip("() "))
            body = prefix.group(3)
        lm2 = re.match(r"(?is)^\s*LOAD\b(.*)$", body)
        is_sql = bool(re.match(r"(?is)^\s*SQL\b", body)) or bool(re.match(r"(?is)^\s*SELECT\b", body))
        if not lm2 and not is_sql:
            continue
        if not label and join_target:
            label = join_target

        kind, source, files, external = "unknown", "", set(), False
        out_fields, wildcard = [], False

        if lm2:
            rest = lm2.group(1)
            km = _SRC_KW.search(rest)
            field_part = rest[:km.start()] if km else rest
            if "*" in _split_top_level(field_part)[0] and re.search(r"(?<![\w])\*", field_part):
                # LOAD *  (all fields from the source)
                wildcard = True
            for piece in _split_top_level(field_part):
                p = piece.strip()
                if not p or p == "*":
                    continue
                out, inn = _field_in_out_name(p)
                if out:
                    out_fields.append({"out": out, "in": inn})
            if km:
                kw = km.group(1).lower()
                tail = rest[km.end():].strip()
                if kw == "from":
                    ref = re.split(r"(?is)\bwhere\b|\bwhile\b", tail)[0]
                    source = " ".join(ref.split())[:300]
                    files = {b for b in extract_file_refs(ref)}
                    kind, external = "from", True
                elif kw == "resident":
                    source = _clean_ident(re.split(r"(?is)\bwhere\b|\border\b", tail)[0])
                    kind, external = "resident", False
                elif kw == "inline":
                    source = "INLINE data"
                    kind, external = "inline", False
                elif kw == "autogenerate":
                    source = "AUTOGENERATE"
                    kind, external = "autogenerate", False
            else:
                # LOAD ... ;  with no source - maybe a preceding load for the
                # next SQL SELECT statement
                nxt = stmts[idx + 1] if idx + 1 < len(stmts) else ""
                ncm = re.search(r"(?is)\bCONNECT\s+TO\s+('([^']+)'|\[([^\]]+)\]|\"([^\"]+)\"|([^\s;(]+))", nxt)
                ncon = (next((g for g in ncm.groups()[1:] if g), "") if ncm else "") or cur_conn
                if re.search(r"(?is)\b(SQL\b|SELECT\b)", nxt):
                    kind, external = "sql", True
                    source = _sql_source_label(nxt, ncon)
        elif is_sql:
            kind, external = "sql", True
            source = _sql_source_label(body, cur_conn)

        tables.append({
            "label": label or (f"(table #{len(tables) + 1})"),
            "out_fields": out_fields,
            "out_lc": {f["out"].lower() for f in out_fields},
            "wildcard": wildcard,
            "kind": kind, "source": source, "files": files,
            "external": external, "raw": " ".join(st.split())[:400],
        })
    return tables


def _origin_kind_label(t):
    """Human label for the origin kind of a parsed table."""
    if t["kind"] == "from":
        q = sorted(b for b in t["files"] if b.endswith(".qvd"))
        if q:
            return "QVD", q[0]
        other = sorted(t["files"])
        if other:
            return "File", other[0]
        return "External source", t["source"] or "FROM"
    if t["kind"] == "sql":
        return "Database", t["source"] or "SQL SELECT"
    if t["kind"] == "inline":
        return "Inline", t["label"]
    if t["kind"] == "autogenerate":
        return "Generated", t["label"]
    return "Source", t["source"] or t["label"]


def trace_field_pipeline(field_name, tables):
    """Follow a field back through the load script to its origin in THIS app.
    Returns {field, found, steps:[...], origin, external_sources:set, note}.
    steps go origin-first; the final consumer (the app) is added by the caller."""
    fl = (field_name or "").strip().lower()

    def producers_of(name_lc, allow_wild=True):
        exact = [t for t in tables if name_lc in t["out_lc"]]
        if exact:
            return exact, False
        if allow_wild:
            wild = [t for t in tables if t["wildcard"]]
            return wild, True
        return [], False

    prod, via_wild = producers_of(fl)
    if not prod:
        return {"field": field_name, "found": False, "steps": [],
                "external_sources": set(), "origin": None, "source_field": field_name,
                "note": "This field is not produced by any LOAD in the script "
                        "(it may be created by a wildcard load, an expression, or "
                        "renamed in a way the text scan could not follow)."}

    steps = []
    external_sources = set()
    seen_tables = set()
    cur = prod[0]
    cur_name = fl
    source_field = field_name   # the field's name in its ORIGIN (qvd/db column)
    note = ""
    if via_wild:
        note = "matched via a wildcard LOAD * - the field name is assumed to pass through unchanged."
    guard = 0
    while cur is not None and guard < 50:
        guard += 1
        steps.append({"table": cur["label"], "kind": cur["kind"],
                      "source": cur["source"], "files": sorted(cur["files"])})
        if cur["label"].lower() in seen_tables:
            note = (note + " " if note else "") + "stopped on a circular resident reference."
            break
        seen_tables.add(cur["label"].lower())
        if cur["kind"] == "resident":
            # what is this field called inside the resident source table?
            in_name = cur_name
            for f in cur["out_fields"]:
                if f["out"].lower() == cur_name and f["in"]:
                    in_name = f["in"].lower()
                    break
            target = next((t for t in tables
                           if t["label"].lower() == cur["source"].lower()), None)
            if target and (in_name in target["out_lc"] or target["wildcard"]):
                cur = target
                cur_name = in_name
                continue
            # resident target not found or doesn't expose the field -> internal origin
            note = (note + " " if note else "") + \
                "the chain ends at an in-app (RESIDENT) table; no external FROM was found for this field."
            break
        # from / sql / inline / autogenerate / unknown -> terminal origin.
        # resolve the field's name in the origin (the qvd/db column it was read as)
        source_field = cur_name
        for f in cur["out_fields"]:
            if f["out"].lower() == cur_name and f["in"]:
                source_field = f["in"]
                break
        for b in cur["files"]:
            external_sources.add(b)
        break

    origin = steps[-1] if steps else None
    steps_origin_first = list(reversed(steps))
    return {"field": field_name, "found": True, "steps": steps_origin_first,
            "origin": origin, "external_sources": external_sources,
            "source_field": source_field, "note": note}


def native_upstream_for_sources(graph, basenames, app_guid=None):
    """Anchor the native lineage walk on the resolved SOURCE node(s) (e.g. the
    QVD the field was loaded from) instead of the whole app, so the cross-app
    chain is slim. Returns a list of resource steps furthest-source first:
    [{kind,label,depth}]. Empty if nothing matched."""
    nodes = graph.get("nodes", {})
    edges = graph.get("edges", [])
    if not nodes or not basenames:
        return []
    bl = {b.lower() for b in basenames}
    preds = {}
    for e in edges:
        s, t = e.get("source"), e.get("target")
        if s and t:
            preds.setdefault(t, []).append(s)

    def node_basenames(qri, meta):
        names = set()
        fp = meta.get("filePath") or ""
        if fp:
            names.add(os.path.basename(fp).lower())
        lab = (meta.get("label") or "").strip().lower()
        if lab:
            names.add(os.path.basename(lab))
            names.add(lab)
        names.add(os.path.basename(str(qri)).lower())
        return names

    starts = [q for q, m in nodes.items() if node_basenames(q, m) & bl]
    if not starts:
        return []

    seen, order = set(), []
    frontier = [(0, q) for q in starts]
    while frontier:
        depth, q = frontier.pop(0)
        if q in seen:
            continue
        seen.add(q)
        m = nodes.get(q, {})
        order.append({"qri": q, "label": m.get("label", "") or q,
                      "kind": _native_node_kind(q, m), "depth": depth})
        for nxt in preds.get(q, []):
            if nxt not in seen:
                frontier.append((depth + 1, nxt))

    keep = ("Database", "Qlik app", "File", "Dataset", "Table")
    res, seenlab = [], set()
    for c in order:
        # skip the depth-0 source file itself - the script pipeline already
        # shows that QVD/file as its [QVD]/[File] origin node (no duplicate).
        if c["depth"] == 0 and c["kind"].startswith(("File", "Dataset")):
            continue
        if c["kind"].startswith(keep) and c["label"] not in seenlab:
            seenlab.add(c["label"])
            res.append(c)
    res.sort(key=lambda x: -x["depth"])
    return res


def render_field_pipeline_text(pipe, app_title, upstream=None, nodes=None):
    """Plain 'pipeline' view: origin -> ... -> producing table -> this report.
    If `nodes` (a precomputed cross-app node list) is given it is used as-is;
    otherwise the single-app node list is built from `pipe`+`upstream`."""
    field = pipe.get("field", "")
    if not pipe.get("found") and not nodes:
        L = [f"DATA PIPELINE   -   {field}     (report: {app_title})", "",
             "  " + (pipe.get("note") or "Could not resolve this field in the load script.")]
        return "\n".join(L)
    if nodes is None:
        nodes = build_pipeline_nodes(pipe, app_title, upstream)
    crossed = any(n.get("kind", "").startswith("Qlik app") for n in nodes)
    L = [f"DATA PIPELINE   -   {field}     (report: {app_title})", "",
         "How this field reached the report (origin first):", ""]
    for i, n in enumerate(nodes):
        lab = "This report" if n["kind"] == "Report" else n["label"]
        prefix = "[This report]" if n["kind"] == "Report" else f"[{n['kind']}]"
        L.append(f"  {prefix}  {lab}")
        if n.get("meta"):
            L.append(f"        {n['meta']}")
        if i < len(nodes) - 1:
            L.append("     |")
    if crossed:
        L += ["", "(apps above this report's own source were resolved through "
                  "Qlik's data lineage + their load scripts)"]
    if pipe.get("note"):
        L += ["", "Note: " + pipe["note"]]
    L += ["", "Read from the load script of each app (best-effort text parse)."]
    return "\n".join(L)


def build_pipeline_nodes(pipe, app_title, upstream=None):
    """Flatten a pipeline into a single top->bottom column of graph nodes
    [{kind,label}] for the SVG renderer. Origin at top, report at bottom."""
    nodes = []
    for c in (upstream or []):
        nodes.append({"kind": c["kind"], "label": c["label"]})
    seen = {(n["kind"], n["label"]) for n in nodes}
    for st in pipe.get("steps", []):
        if st["kind"] == "from":
            k, lab = _origin_kind_label(st)
            key = (k, lab)
            if key not in seen:
                seen.add(key)
                nodes.append({"kind": k, "label": lab})
        elif st["kind"] == "sql":
            nodes.append({"kind": "Database", "label": st["source"] or "SQL SELECT"})
        elif st["kind"] == "inline":
            nodes.append({"kind": "Inline", "label": "INLINE data"})
        elif st["kind"] == "autogenerate":
            nodes.append({"kind": "Generated", "label": "AUTOGENERATE"})
        nodes.append({"kind": "Table", "label": st["table"]})
    nodes.append({"kind": "Report", "label": app_title})
    return nodes


def write_field_pipeline_graph_html(pipe, app_title, app_guid, out_dir, upstream=None, nodes=None):
    """Render the field pipeline as a clean, single-column vertical graph
    (origin at top -> report at bottom). Always small and readable.
    If `nodes` (a precomputed cross-app node list, items may carry 'meta') is
    given it is used as-is; otherwise it is built from `pipe`+`upstream`."""
    os.makedirs(out_dir, exist_ok=True)
    stamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    fname = f"pipeline_{safe(pipe.get('field', 'field'))}_{safe(app_title)}_{safe(app_guid)}_{stamp}.html"
    path = os.path.join(out_dir, fname)

    def esc(s):
        return html.escape(str(s or ""))

    if nodes is None:
        nodes = build_pipeline_nodes(pipe, app_title, upstream)

    has_meta = any(n.get("meta") for n in nodes)
    BW, BH, VG, MX, MY = 360, (68 if has_meta else 56), 40, 40, 30
    width = MX * 2 + BW
    height = MY * 2 + len(nodes) * BH + (len(nodes) - 1) * VG

    palette = [("Database", "#7a5230"), ("QVD", "#3E7185"), ("File", "#5B6B72"),
               ("Qlik app", "#315C6D"), ("Dataset", "#3E7185"), ("Table", "#4B7A8C"),
               ("Inline", "#5B6B72"), ("Generated", "#5B6B72"), ("Report", "#274A57"),
               ("External source", "#7a5230"), ("Source", "#5B6B72")]

    def color(kind):
        for k, c in palette:
            if kind.startswith(k):
                return c
        return "#5B6B72"

    svg = [f'<svg width="{width}" height="{height}" xmlns="http://www.w3.org/2000/svg" '
           'font-family="Segoe UI, Arial, sans-serif">',
           '<defs><marker id="arrow" markerWidth="10" markerHeight="10" refX="8" refY="3" '
           'orient="auto" markerUnits="strokeWidth"><path d="M0,0 L8,3 L0,6 Z" fill="#9aa6ab"/></marker></defs>']
    x = MX
    for i in range(len(nodes) - 1):
        y1 = MY + i * (BH + VG) + BH
        y2 = MY + (i + 1) * (BH + VG)
        cx = x + BW / 2
        svg.append(f'<line x1="{cx:.0f}" y1="{y1}" x2="{cx:.0f}" y2="{y2}" '
                   'stroke="#9aa6ab" stroke-width="1.5" marker-end="url(#arrow)"/>')
    for i, n in enumerate(nodes):
        y = MY + i * (BH + VG)
        lab = n["label"]
        lab = (lab[:42] + "…") if len(lab) > 42 else lab
        meta = n.get("meta") or ""
        meta = (meta[:54] + "…") if len(meta) > 54 else meta
        svg.append(f'<rect x="{x}" y="{y}" width="{BW}" height="{BH}" rx="10" fill="{color(n["kind"])}"/>')
        svg.append(f'<text x="{x + BW / 2:.0f}" y="{y + 18}" fill="#cfe0e6" font-size="10" '
                   f'text-anchor="middle">{esc(n["kind"])}</text>')
        svg.append(f'<text x="{x + BW / 2:.0f}" y="{y + 37}" fill="#ffffff" font-size="13" '
                   f'font-weight="600" text-anchor="middle">{esc(lab)}</text>')
        if meta:
            svg.append(f'<text x="{x + BW / 2:.0f}" y="{y + 55}" fill="#cfe0e6" font-size="9" '
                       f'text-anchor="middle">{esc(meta)}</text>')
    svg.append('</svg>')

    note = pipe.get("note") or ""
    page = ['<!doctype html><meta charset="utf-8">',
            '<style>body{font-family:Segoe UI,Arial,sans-serif;color:#1F2A30;margin:24px;background:#EEF1F3;}'
            'h1{color:#315C6D;} .muted{color:#5B6B72;} .wrap{background:#fff;border:1px solid #D5DCDF;'
            'border-radius:10px;padding:16px;display:inline-block;}</style>',
            f"<h1>How does &ldquo;{esc(pipe.get('field', ''))}&rdquo; reach this report?</h1>",
            f'<p class="muted">Report: {esc(app_title)} &middot; flow reads top (origin) to bottom (report)</p>',
            '<div class="wrap">', "".join(svg), '</div>']
    if note:
        page.append(f'<p class="muted" style="margin-top:14px;max-width:600px;">Note: {esc(note)}</p>')
    page.append('<p class="muted" style="margin-top:6px;">Read from this app’s load script '
                '(best-effort) and Qlik’s data lineage for the upstream apps.</p>')
    open(path, "w", encoding="utf-8").write("\n".join(page))
    return path


# ============================================================
#  Cross-app pipeline: hop into the apps that PRODUCE a source
#  QVD and parse THEIR load script to resolve the real origin
#  (the database/file), plus app metadata (space/owner/reload).
# ============================================================
def _rest_get(tenant, api_key, path, timeout=30):
    host = normalize_host(tenant)
    url = f"https://{host}{path}"
    req = urllib.request.Request(url, headers={"Authorization": "Bearer " + api_key})
    with urllib.request.urlopen(req, timeout=timeout) as resp:
        return json.loads(resp.read().decode("utf-8"))


def _guid_from_qri(qri):
    m = re.search(r"sense://([0-9a-fA-F-]{36})", str(qri or ""))
    return m.group(1) if m else ""


def native_producer_apps(graph, basenames):
    """For the given source QVD basename(s), return the Qlik app(s) that
    PRODUCE them (STORE ... INTO that qvd), discovered from the lineage graph.
    Returns [{guid, label, qvd}]. Walks predecessors past PROCESSOR nodes
    until it reaches an app node."""
    nodes = graph.get("nodes", {})
    edges = graph.get("edges", [])
    if not nodes or not basenames:
        return []
    bl = {b.lower() for b in basenames}
    preds = {}
    for e in edges:
        s, t = e.get("source"), e.get("target")
        if s and t:
            preds.setdefault(t, []).append(s)

    def node_basenames(qri, meta):
        names = set()
        fp = meta.get("filePath") or ""
        if fp:
            names.add(os.path.basename(fp).lower())
        lab = (meta.get("label") or "").strip().lower()
        if lab:
            names.add(os.path.basename(lab))
            names.add(lab)
        names.add(os.path.basename(str(qri)).lower())
        return names

    qvd_nodes = [(q, [b for b in node_basenames(q, m) if b in bl])
                 for q, m in nodes.items() if node_basenames(q, m) & bl]
    out, seen = [], set()
    for qnode, matched in qvd_nodes:
        qvd_name = matched[0] if matched else qnode
        frontier = list(preds.get(qnode, []))
        local_seen = set()
        while frontier:
            p = frontier.pop(0)
            if p in local_seen:
                continue
            local_seen.add(p)
            m = nodes.get(p, {})
            if _native_node_kind(p, m).startswith("Qlik app"):
                g = _guid_from_qri(p)
                if g and g not in seen:
                    seen.add(g)
                    out.append({"guid": g, "label": m.get("label", "") or g, "qvd": qvd_name})
                continue  # stop at the producing app
            frontier.extend(preds.get(p, []))
    return out


def pipe_origin_node(pipe):
    """The single ORIGIN node ({kind,label}) of a pipe (its furthest step)."""
    steps = pipe.get("steps", []) if pipe else []
    if not steps:
        return None
    st = steps[0]   # origin-first
    if st["kind"] == "from":
        k, lab = _origin_kind_label(st)
        return {"kind": k, "label": lab}
    if st["kind"] == "sql":
        return {"kind": "Database", "label": st["source"] or "SQL SELECT"}
    if st["kind"] == "inline":
        return {"kind": "Inline", "label": "INLINE data"}
    if st["kind"] == "autogenerate":
        return {"kind": "Generated", "label": "AUTOGENERATE"}
    return {"kind": "Table", "label": st["table"]}


def app_meta(tenant, api_key, guid, space_cache=None, user_cache=None):
    """Fetch {space, owner, reload} for an app via REST. Caches space/user
    lookups. Best-effort: missing pieces come back blank."""
    space_cache = space_cache if space_cache is not None else {}
    user_cache = user_cache if user_cache is not None else {}
    meta = {"space": "", "owner": "", "reload": ""}
    try:
        attrs = _rest_get(tenant, api_key, f"/api/v1/apps/{guid}").get("attributes", {})
    except Exception:
        return meta
    meta["reload"] = attrs.get("lastReloadTime", "") or ""
    sid = attrs.get("spaceId") or ""
    if not sid:
        meta["space"] = "Personal"
    elif sid in space_cache:
        meta["space"] = space_cache[sid]
    else:
        try:
            nm = _rest_get(tenant, api_key, f"/api/v1/spaces/{sid}").get("name", sid)
        except Exception:
            nm = sid
        space_cache[sid] = nm
        meta["space"] = nm
    uid = attrs.get("ownerId") or ""
    if uid:
        if uid in user_cache:
            meta["owner"] = user_cache[uid]
        else:
            try:
                u = _rest_get(tenant, api_key, f"/api/v1/users/{uid}")
                nm = u.get("name", "") or (u.get("attributes", {}) or {}).get("name", "")
            except Exception:
                nm = ""
            user_cache[uid] = nm
            meta["owner"] = nm
    return meta


def _meta_str(meta):
    if not meta:
        return ""
    bits = []
    if meta.get("space"):
        bits.append("Space: " + meta["space"])
    if meta.get("owner"):
        bits.append("Owner: " + meta["owner"])
    if meta.get("reload"):
        bits.append("Reloaded: " + str(meta["reload"]).replace("T", " ").replace("Z", " UTC")[:19])
    return "  ·  ".join(bits)


def trace_upstream_apps(graph, start_sources, start_field, open_app,
                        meta_fn=None, log=None, max_hops=6):
    """Walk app-by-app: for each source QVD, find the producing app (from the
    lineage graph), open it, parse ITS load script and trace the field to its
    own origin (db/file/another qvd). Recurse on qvd origins.

    open_app(guid) -> {"title","script","reload"} or None
    meta_fn(guid)  -> {"space","owner","reload"} or {}

    Returns hops nearest-first:
      [{guid,title,pipe,meta,produced_qvd,depth}]"""
    hops, seen = [], set()
    frontier = [(1, q, start_field) for q in sorted(start_sources) if q.endswith(".qvd")]
    while frontier:
        depth, qvd, fld = frontier.pop(0)
        if depth > max_hops:
            continue
        for prod in native_producer_apps(graph, {qvd}):
            g = prod["guid"]
            if g in seen:
                continue
            seen.add(g)
            info = open_app(g) if open_app else None
            meta = (meta_fn(g) if meta_fn else {}) or {}
            if not info:
                if log:
                    log(f"  (upstream app {prod['label']} could not be opened - "
                        "showing it without its source detail)")
                hops.append({"guid": g, "title": prod["label"], "pipe": None,
                             "meta": meta, "produced_qvd": qvd, "depth": depth})
                continue
            if info.get("reload") and not meta.get("reload"):
                meta["reload"] = info["reload"]
            tables = parse_load_tables(info.get("script", ""))
            pipe = trace_field_pipeline(fld, tables)
            hops.append({"guid": g, "title": info.get("title", prod["label"]),
                         "pipe": pipe, "meta": meta, "produced_qvd": qvd, "depth": depth})
            if log:
                src = sorted(pipe.get("external_sources") or [])
                log(f"  upstream: {info.get('title', g)} -> "
                    f"{', '.join(src) if src else 'in-app/db origin'}")
            if pipe.get("found"):
                nf = pipe.get("source_field") or fld
                for nq in sorted(pipe.get("external_sources") or []):
                    if nq.endswith(".qvd"):
                        frontier.append((depth + 1, nq, nf))
    hops.sort(key=lambda h: h["depth"])
    return hops


def assemble_pipeline_nodes(this_pipe, this_app_title, this_meta, hops):
    """Build the full origin-first node list across apps. Each upstream hop
    contributes its own ORIGIN (db/file, resolved from that app's script) then
    its [Qlik app] node; the handoff QVDs and this app's tables follow."""
    nodes = []

    def push(n):
        if nodes and nodes[-1].get("kind") == n.get("kind") and \
           nodes[-1].get("label") == n.get("label"):
            return  # collapse an adjacent duplicate (e.g. repeated QVD)
        nodes.append(n)

    for hop in sorted(hops, key=lambda h: -h["depth"]):   # furthest first
        onode = pipe_origin_node(hop.get("pipe"))
        if onode:
            push(onode)
        push({"kind": "Qlik app", "label": hop["title"], "meta": _meta_str(hop.get("meta"))})
    # nearest: this report's own pipeline (origin-first: handoff qvd -> tables)
    for st in this_pipe.get("steps", []):
        if st["kind"] == "from":
            k, lab = _origin_kind_label(st)
            push({"kind": k, "label": lab})
        elif st["kind"] == "sql":
            push({"kind": "Database", "label": st["source"] or "SQL SELECT"})
        elif st["kind"] == "inline":
            push({"kind": "Inline", "label": "INLINE data"})
        elif st["kind"] == "autogenerate":
            push({"kind": "Generated", "label": "AUTOGENERATE"})
        push({"kind": "Table", "label": st["table"]})
    push({"kind": "Report", "label": this_app_title, "meta": _meta_str(this_meta)})
    return nodes
