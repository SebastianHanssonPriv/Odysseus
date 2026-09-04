"""Per-table source resolution for Power BI semantic models (datasets).

Combines the scanner API's per-table M expression with the dataflow export
API's per-entity M code to answer, for each table in each semantic model,
which warehouse table/view (or other data source) it ultimately reads from,
and which of that source's fields it keeps -- chasing through Gen1 dataflow
references when a table's own source is "Get Data > Power BI dataflows"
rather than a direct connector call, and through same-document staging-query
references at both the dataset and the dataflow level.
"""

from __future__ import annotations

import csv
import datetime
import os
import re
from dataclasses import asdict, dataclass, field

from dataflow_admin import export_dataflow
from mashup_parser import MAX_REFERENCE_DEPTH, resolve_source, split_shared_queries
from scanner import list_workspace_ids, scan_workspaces

# Dataflow-boundary hops (dataset -> dataflow -> dataflow -> ...) are capped
# independently of mashup_parser's own same-document reference-chasing cap,
# since a dataflow can itself be built on another dataflow (linked entities).
_MAX_DATAFLOW_HOPS = MAX_REFERENCE_DEPTH


@dataclass
class TableSourceResult:
    workspace_id: str
    workspace_name: str
    dataset_id: str
    dataset_name: str
    table_name: str
    status: str
    hops: list[dict] = field(default_factory=list)
    direct_sources: list[dict] = field(default_factory=list)
    column_usage: list[dict] = field(default_factory=list)
    note: str | None = None


class DataflowCache:
    """Fetches and parses each dataflow's export at most once per run."""

    def __init__(self, client):
        self._client = client
        self._queries: dict[str, dict[str, str] | None] = {}  # None = export/parse failed
        self._errors: dict[str, str] = {}

    def queries_for(self, dataflow_id: str) -> dict[str, str] | None:
        if dataflow_id not in self._queries:
            try:
                definition = export_dataflow(self._client, dataflow_id)
                document = definition.get("pbi:mashup", {}).get("document", "")
                if not document:
                    self._queries[dataflow_id] = None
                    self._errors[dataflow_id] = (
                        "export succeeded but no pbi:mashup.document found in the "
                        "response -- unexpected/unverified shape, see "
                        "dataflow_admin.export_dataflow's docstring"
                    )
                else:
                    self._queries[dataflow_id] = split_shared_queries(document)
            except Exception as exc:  # noqa: BLE001 - one dataflow's failure must not abort the run
                self._queries[dataflow_id] = None
                self._errors[dataflow_id] = str(exc)
        return self._queries[dataflow_id]

    def error_for(self, dataflow_id: str) -> str | None:
        return self._errors.get(dataflow_id)


def resolve_table_source(
    workspace_id: str,
    workspace_name: str,
    dataset_id: str,
    dataset_name: str,
    table_name: str,
    expression: str,
    dataset_siblings: dict[str, str],
    dataflow_cache: DataflowCache,
) -> TableSourceResult:
    base = dict(
        workspace_id=workspace_id,
        workspace_name=workspace_name,
        dataset_id=dataset_id,
        dataset_name=dataset_name,
        table_name=table_name,
    )
    hops: list[dict] = []
    current_expr = expression
    current_siblings = dataset_siblings
    current_label = f"dataset:{dataset_name}.{table_name}"

    for _ in range(_MAX_DATAFLOW_HOPS + 1):
        resolution = resolve_source(current_expr, sibling_queries=current_siblings)

        if resolution.status == "dataflow_reference":
            ref = resolution.dataflow_ref
            hops.append(
                {
                    "from": current_label,
                    "dataflow_workspace_id": ref.workspace_id,
                    "dataflow_id": ref.dataflow_id,
                    "entity": ref.entity,
                    "same_document_hops": resolution.reference_chain,
                }
            )
            if not ref.dataflow_id or not ref.entity:
                return TableSourceResult(
                    **base,
                    status="dataflow_reference_incomplete",
                    hops=hops,
                    note="dataflow reference found but workspaceId/dataflowId/entity could not all be extracted from the M code",
                )

            queries = dataflow_cache.queries_for(ref.dataflow_id)
            if queries is None:
                return TableSourceResult(
                    **base,
                    status="dataflow_export_failed",
                    hops=hops,
                    note=dataflow_cache.error_for(ref.dataflow_id),
                )
            if ref.entity not in queries:
                return TableSourceResult(
                    **base,
                    status="dataflow_entity_not_found",
                    hops=hops,
                    note=f"entity '{ref.entity}' not found among {len(queries)} parsed queries in dataflow {ref.dataflow_id}",
                )

            current_expr = queries[ref.entity]
            current_siblings = queries
            current_label = f"dataflow:{ref.dataflow_id}.{ref.entity}"
            continue

        if resolution.status in ("direct_source", "multiple_direct_sources"):
            return TableSourceResult(
                **base,
                status=resolution.status,
                hops=hops,
                direct_sources=[asdict(s) for s in resolution.direct_sources],
            )

        return TableSourceResult(**base, status="unresolved", hops=hops, note=resolution.note)

    return TableSourceResult(
        **base,
        status="max_hops_exceeded",
        hops=hops,
        note=f"dataflow reference chain exceeded {_MAX_DATAFLOW_HOPS} hops",
    )


# ============================================================
#  Tenant-wide batch scan + report (the `model-lineage` feature)
# ============================================================
_NO_EXPRESSION_NOTE = (
    "Scanner API returned no M expression for this table. Most likely cause: "
    "the tenant setting 'Enhance admin APIs responses with DAX and mashup "
    "expressions' is not enabled in Admin portal -> Tenant settings."
)
_NO_TABLES_NOTE = (
    "Scanner API returned this dataset with an empty/missing 'tables' list (no "
    "per-table detail at all, not even a name without M code). Most likely "
    "cause: the tenant setting 'Enhance admin APIs responses with detailed "
    "metadata' is not enabled in Admin portal -> Tenant settings, or is not "
    "enabled for this service principal's security group specifically."
)

# Statuses whose row gets a warning tint in the Excel report.
_BAD_STATUSES = {"unresolved", "no_expression_available", "dataflow_export_failed",
                 "dataflow_entity_not_found", "dataset_has_no_tables"}
_WARN_STATUSES = {"dataflow_reference_incomplete", "max_hops_exceeded", "multiple_direct_sources"}


def scan_model_lineage(client, scan_timeout_seconds=600, cancel_check=None, log=print):
    """Tenant-wide: for every table in every semantic model, resolve its
    source (direct connector, or chased through Gen1 dataflow(s)) and, where
    the M code says so explicitly, which fields survive. Returns a list of
    TableSourceResult as dicts. cancel_check, if given, is polled between
    workspaces and stops the scan early without raising."""
    dataflow_cache = DataflowCache(client)
    workspace_ids = list(list_workspace_ids(client))
    results = []
    batch_errors = 0
    tally = {"datasets": 0, "reports": 0, "dataflows": 0, "dashboards": 0}

    for workspace in scan_workspaces(client, workspace_ids, scan_timeout_seconds):
        if cancel_check and cancel_check():
            log("Model lineage scan cancelled.")
            break
        if "scan_batch_error" in workspace:
            batch_errors += 1
            log(f"  scan batch error: {workspace['scan_batch_error']}")
            continue
        for key in tally:
            tally[key] += len(workspace.get(key) or [])
        results.extend(_resolve_workspace_datasets(workspace, dataflow_cache, log))

    real_tables = [r for r in results if r["status"] != "dataset_has_no_tables"]
    log(f"Scanned {len(workspace_ids)} workspace(s) ({batch_errors} batch error(s)), "
        f"resolved {len(real_tables)} table(s) across {len(results) - len(real_tables)} "
        f"dataset(s) with no table detail.")
    if not real_tables:
        all_empty = tally["datasets"] == tally["reports"] == tally["dataflows"] == tally["dashboards"] == 0
        log(f"  Diagnostic: across all scanned workspaces the Scanner API returned "
            f"{tally['datasets']} dataset(s), {tally['reports']} report(s), "
            f"{tally['dataflows']} dataflow(s), {tally['dashboards']} dashboard(s). " +
            ("All zero -> the service principal likely cannot see workspace content at all "
             "(check it is in the Power BI admin security group and 'Allow service principals "
             "to use read-only admin APIs' is enabled for that group)."
             if all_empty else
             "Datasets/reports/dataflows are present but every dataset came back with an "
             "empty 'tables' list -> almost certainly the tenant setting 'Enhance admin APIs "
             "responses with detailed metadata' is not enabled for this service principal's "
             "security group (Admin portal -> Tenant settings)."))
    return results


_DAX_BRACKET_RE = re.compile(r"\[([^\]]+)\]")


def dax_referenced_fields(tables):
    """Every column name referenced by any calculated column or measure's
    DAX expression anywhere in a dataset's tables (DAX allows cross-table
    references, e.g. 'Orders'[Amount], so this is scoped to the whole
    dataset, not just one table). Returns a lowercased set.

    This is the closest signal available from Power BI's Admin APIs to
    "used in a report": unlike Qlik's Engine API, Power BI's Admin APIs
    expose no visual/report-page content, so a raw column placed directly on
    a visual with no calculation involved cannot be detected at all -- a
    'not referenced' column is a candidate to verify, not a verdict. Same
    text-matching caveat as Qlik's Usage analysis: a column named only
    inside a string literal, or via a name built from a DAX variable, would
    be missed or falsely matched -- this is a scan, not an evaluator."""
    refs = set()
    for t in tables:
        for col in (t.get("columns") or []):
            expr = col.get("expression")
            if expr:
                refs |= {m.group(1).strip().lower() for m in _DAX_BRACKET_RE.finditer(expr)}
        for m_ in (t.get("measures") or []):
            expr = m_.get("expression")
            if expr:
                refs |= {m.group(1).strip().lower() for m in _DAX_BRACKET_RE.finditer(expr)}
    return refs


def _column_usage_for_table(table, dax_refs):
    """Every column the Scanner API lists on this model table (from dataset
    schema detail, independent of whether this table's own warehouse source
    could be resolved), and whether it is DAX-referenced anywhere in the
    dataset."""
    return [{"column": c.get("name", ""), "used_in_dax": (c.get("name") or "").lower() in dax_refs}
           for c in (table.get("columns") or []) if c.get("name")]


def _resolve_workspace_datasets(workspace, dataflow_cache, log):
    workspace_id = workspace.get("id", "")
    workspace_name = workspace.get("name", "")
    results = []

    for dataset in workspace.get("datasets", []):
        dataset_id = dataset.get("id", "")
        dataset_name = dataset.get("name", "")
        tables = dataset.get("tables", [])
        if not tables:
            results.append(asdict(TableSourceResult(
                workspace_id=workspace_id, workspace_name=workspace_name,
                dataset_id=dataset_id, dataset_name=dataset_name, table_name="",
                status="dataset_has_no_tables", note=_NO_TABLES_NOTE,
            )))
            continue
        dataset_siblings = _dataset_sibling_expressions(tables)
        dax_refs = dax_referenced_fields(tables)

        for table in tables:
            table_name = table.get("name", "")
            expression = _table_expression(table)
            if expression is None:
                result = TableSourceResult(
                    workspace_id=workspace_id, workspace_name=workspace_name,
                    dataset_id=dataset_id, dataset_name=dataset_name, table_name=table_name,
                    status="no_expression_available", note=_NO_EXPRESSION_NOTE,
                )
            else:
                result = resolve_table_source(
                    workspace_id, workspace_name, dataset_id, dataset_name, table_name,
                    expression, dataset_siblings, dataflow_cache,
                )
            result.column_usage = _column_usage_for_table(table, dax_refs)
            results.append(asdict(result))
        if tables:
            log(f"  {workspace_name} / {dataset_name}: {len(tables)} table(s) resolved")
    return results


def _table_expression(table):
    # Documented Scanner API shape: table.source -> [{"expression": "..."}].
    source = table.get("source")
    if not isinstance(source, list) or not source:
        return None
    first = source[0]
    return first.get("expression") if isinstance(first, dict) else None


def _dataset_sibling_expressions(tables):
    siblings = {}
    for t in tables:
        expr = _table_expression(t)
        name = t.get("name")
        if name and expr:
            siblings[name] = expr
    return siblings


def render_model_lineage_text(results):
    counts: dict[str, int] = {}
    for r in results:
        counts[r["status"]] = counts.get(r["status"], 0) + 1
    all_cols = [cu for r in results for cu in (r.get("column_usage") or [])]
    used_cols = sum(1 for cu in all_cols if cu["used_in_dax"])
    lines = [f"MODEL LINEAGE - {len(results)} table(s) scanned", ""]
    for status, n in sorted(counts.items(), key=lambda kv: -kv[1]):
        lines.append(f"  {status}: {n}")
    if all_cols:
        lines.append(f"  Columns: {len(all_cols)} total, {used_cols} referenced by a DAX "
                     f"calculation (measure or calculated column) somewhere in their dataset")
    lines.append("")
    flagged = [r for r in results if r["status"] in _BAD_STATUSES]
    for r in flagged[:20]:
        lines.append(f"  {r['status'].upper()}: {r['workspace_name']} / {r['dataset_name']} / {r['table_name']}")
    if len(flagged) > 20:
        lines.append(f"  ... and {len(flagged) - 20} more - see the Excel report.")
    lines.append("")
    lines.append("Full detail, including every resolved source table, field, and column usage, "
                 "is in the Excel report.")
    return "\n".join(lines)


def _summarize_row(record):
    direct = record.get("direct_sources") or []
    connectors = "; ".join(d["connector"] for d in direct)
    resolved_tables = "; ".join(d["resolved_table"] for d in direct if d.get("resolved_table"))
    fields, field_sources = [], set()
    for d in direct:
        if d.get("fields"):
            fields.extend(d["fields"])
            field_sources.add(d.get("fields_source") or "")
    return [
        record["workspace_name"], record["dataset_name"], record["table_name"],
        record["status"], len(record.get("hops") or []), connectors, resolved_tables,
        ", ".join(fields), "/".join(sorted(field_sources)), record.get("note") or "",
    ]


def write_model_lineage_report(results, out_dir, log):
    """One combined workbook: Summary (status counts) + Model lineage detail
    -- one row per table: workspace, dataset, table, status, dataflow hop
    count, connector, source table/view, and fields where the M code made
    them explicit (see mashup_parser.extract_selected_fields)."""
    os.makedirs(out_dir, exist_ok=True)
    stamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    base = f"model_lineage_{stamp}"

    status_counts: dict[str, int] = {}
    for r in results:
        status_counts[r["status"]] = status_counts.get(r["status"], 0) + 1
    detail_rows = [_summarize_row(r) for r in results]

    column_headers = ["Workspace", "Dataset", "Table", "Column", "Used in a DAX calculation"]
    column_rows = [
        [r["workspace_name"], r["dataset_name"], r["table_name"], cu["column"],
         "Yes" if cu["used_in_dax"] else "No"]
        for r in results for cu in (r.get("column_usage") or [])
    ]
    total_cols = len(column_rows)
    used_cols = sum(1 for r in column_rows if r[-1] == "Yes")

    warn = ("READ FIRST - this is a best-effort text-level scan of each table's Power Query M "
            "code, not an M interpreter (see mashup_parser.PARSER_LIMITATIONS). Fields are only "
            "listed when the M code uses a native SQL Query= passthrough or an explicit "
            "Table.SelectColumns call; a blank Fields column means every source column passes "
            "through unnarrowed by this scan, not that the table has no fields. "
            "'no_expression_available' usually means the tenant setting 'Enhance admin APIs "
            "responses with DAX and mashup expressions' is not enabled. 'Used in a DAX "
            "calculation' (Column usage sheet) means the column is referenced by a measure or "
            "calculated column's DAX expression somewhere in its dataset - Power BI's Admin APIs "
            "expose no report/visual content (unlike Qlik's Engine API), so a raw column placed "
            "directly on a visual with no calculation involved cannot be detected; a 'No' is a "
            "candidate to verify by hand, not a verdict that the column is unused.")

    summary_headers = ["Status", "Table count"]
    summary_rows = sorted(status_counts.items(), key=lambda kv: -kv[1])
    summary_rows.append(("Columns referenced by a DAX calculation",
                         f"{used_cols} / {total_cols}" if total_cols else "0 / 0"))
    detail_headers = ["Workspace", "Dataset", "Table", "Status", "Dataflow hops",
                      "Connector", "Source table/view", "Fields (where resolved)",
                      "Fields detected via", "Note"]

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except Exception:
        detail_path = os.path.join(out_dir, f"{base}.csv")
        with open(detail_path, "w", newline="", encoding="utf-8-sig") as f:
            w = csv.writer(f)
            w.writerow(detail_headers)
            w.writerows(detail_rows)
        column_path = os.path.join(out_dir, f"{base}_column_usage.csv")
        with open(column_path, "w", newline="", encoding="utf-8-sig") as f:
            w = csv.writer(f)
            w.writerow(column_headers)
            w.writerows(column_rows)
        log("openpyxl not installed - wrote CSV files instead of one workbook.")
        return detail_path

    wb = Workbook()
    head_font = Font(bold=True, color="FFFFFF")
    head_fill = PatternFill("solid", fgColor="315C6D")
    warn_font = Font(bold=True, color="B00020")
    wrap = Alignment(wrap_text=True, vertical="top")
    status_fill = {s: PatternFill("solid", fgColor="F7D9DE") for s in _BAD_STATUSES}
    status_fill.update({s: PatternFill("solid", fgColor="F7ECD2") for s in _WARN_STATUSES})

    def style_header_row(ws, row, ncols):
        for c in range(1, ncols + 1):
            ws.cell(row=row, column=c).font = head_font
            ws.cell(row=row, column=c).fill = head_fill

    ws = wb.active
    ws.title = "Summary"
    ws.append(["IMPORTANT - PLEASE READ"])
    ws["A1"].font = warn_font
    ws.append([warn])
    ws.cell(row=ws.max_row, column=1).font = warn_font
    ws.cell(row=ws.max_row, column=1).alignment = wrap
    ws.append([])
    ws.append(summary_headers)
    style_header_row(ws, ws.max_row, len(summary_headers))
    for status, n in summary_rows:
        ws.append([status, n])
    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 14

    ds = wb.create_sheet("Model lineage")
    ds.append(detail_headers)
    style_header_row(ds, 1, len(detail_headers))
    status_col = detail_headers.index("Status")
    for r in detail_rows:
        rn = ds.max_row + 1
        ds.append(r)
        fill = status_fill.get(r[status_col])
        if fill:
            for c in range(1, len(detail_headers) + 1):
                ds.cell(row=rn, column=c).fill = fill
    for i, w in enumerate((22, 22, 22, 26, 12, 20, 26, 40, 16, 40), 1):
        ds.column_dimensions[get_column_letter(i)].width = w
    if detail_rows:
        ds.freeze_panes = "A2"
        ds.auto_filter.ref = f"A1:{get_column_letter(len(detail_headers))}{ds.max_row}"

    bad_fill = PatternFill("solid", fgColor="F7D9DE")
    cu = wb.create_sheet("Column usage")
    cu.append(column_headers)
    style_header_row(cu, 1, len(column_headers))
    used_col_idx = column_headers.index("Used in a DAX calculation")
    for r in column_rows:
        rn = cu.max_row + 1
        cu.append(r)
        if r[used_col_idx] == "No":
            for c in range(1, len(column_headers) + 1):
                cu.cell(row=rn, column=c).fill = bad_fill
    for i, w in enumerate((22, 22, 22, 26, 22), 1):
        cu.column_dimensions[get_column_letter(i)].width = w
    if column_rows:
        cu.freeze_panes = "A2"
        cu.auto_filter.ref = f"A1:{get_column_letter(len(column_headers))}{cu.max_row}"

    out_path = os.path.join(out_dir, f"{base}.xlsx")
    wb.save(out_path)
    return out_path
