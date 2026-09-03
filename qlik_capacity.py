"""qlik_capacity - tenant-wide capacity & redundancy inventory for Qlik Cloud.

Drop-in companion to qlik_core.py. It fetches the data the capacity problem
actually turns on - each app's in-memory data-model SIZE and its per-field /
per-table CARDINALITY - across every app on the tenant, plus a best-effort
data-file (QVD / storage) inventory. Then analyze_capacity() ranks the biggest
savings and redundancy candidates so you know where to cut first.

Reuses qlik_core's host / listing / auth helpers, so it follows the same
conventions already in your tool (Bearer auth, links.next pagination,
best-effort try/except, a `log` callback).

    from qlik_capacity import fetch_capacity_inventory, analyze_capacity, \
                              print_capacity_summary

    inv      = fetch_capacity_inventory(TENANT, API_KEY, log=print)
    findings = analyze_capacity(inv)
    print_capacity_summary(inv, findings)

WHAT IT FETCHES, PER APP
  - space, owner, last reload, created / modified / published   (app attributes)
  - in-memory data-model size (static_byte_size)                (data/metadata)
  - per FIELD: cardinality (distinct values), byte size, system/hidden, tags
  - per TABLE: row count, field count, key fields, byte size
  - best-effort tenant data-file list with sizes                (data-files)

ENDPOINT NOTE  (worth confirming once against your tenant)
  GET /api/v1/apps/{id}/data/metadata  ->  static_byte_size, fields[], tables[]
  This reads the LAST-RELOAD cached model, so an app that has never reloaded
  returns no size - that is captured as an error row, never a crash. Field key
  names are read defensively (.get with fallbacks) in case your tenant's build
  uses slightly different spellings.
"""
import os
import re
import csv
import json
import time
import datetime
import urllib.request
import urllib.error

from qlik_core import (normalize_host, list_apps, list_spaces, QlikExporter,
                       parse_store_reads, extract_file_refs, parse_load_tables)

try:                                  # reuse qlik_core's date parser if present
    from qlik_core import _parse_dt
except Exception:                     # pragma: no cover - tiny local fallback
    def _parse_dt(s):
        if not s:
            return None
        try:
            return datetime.datetime.fromisoformat(str(s).replace("Z", "+00:00"))
        except Exception:
            return None


# ------------------------------------------------------------------ cancellation
class ScanCancelled(Exception):
    """Raised by the long sweeps when a should_cancel() callback returns True, so
    a desktop Cancel button can stop a tenant-wide scan between steps."""


def _ck(should_cancel):
    if should_cancel and should_cancel():
        raise ScanCancelled()


# ------------------------------------------------------------------ REST helper
def _request_json(url, api_key, timeout=30, retries=3):
    """GET a URL with Bearer auth. Retries briefly on 429/503 (rate limits), so a
    big sweep degrades gracefully instead of filling up with throttle errors."""
    req = urllib.request.Request(url, headers={"Authorization": f"Bearer {api_key}"})
    for attempt in range(retries + 1):
        try:
            with urllib.request.urlopen(req, timeout=timeout) as resp:
                return json.loads(resp.read().decode("utf-8"))
        except urllib.error.HTTPError as e:
            if e.code in (429, 503) and attempt < retries:
                ra = e.headers.get("Retry-After") if e.headers else None
                try:
                    wait = float(ra) if ra else 1.5 * (attempt + 1)
                except (TypeError, ValueError):
                    wait = 1.5 * (attempt + 1)
                time.sleep(min(wait, 10))
                continue
            raise
        except urllib.error.URLError:
            # transient network / DNS (getaddrinfo failed, reset, timeout) - retry
            if attempt < retries:
                time.sleep(1.5 * (attempt + 1))
                continue
            raise


def _get_json(tenant, api_key, path, timeout=30):
    """GET a Qlik Cloud REST path and return parsed JSON (429/503-aware)."""
    return _request_json(f"https://{normalize_host(tenant)}{path}", api_key, timeout)


def _get_url(api_key, url, timeout=30):
    """GET an absolute URL (used to follow a links.next href verbatim)."""
    return _request_json(url, api_key, timeout)


def _user_name(tenant, api_key, uid, cache):
    if not uid:
        return ""
    if uid in cache:
        return cache[uid]
    try:
        u = _get_json(tenant, api_key, f"/api/v1/users/{uid}")
        nm = u.get("name", "") or (u.get("attributes", {}) or {}).get("name", "")
    except Exception:
        nm = ""
    cache[uid] = nm
    return nm


def _num(d, *keys):
    """First numeric value among the given keys (tolerates key-name drift)."""
    for k in keys:
        v = d.get(k)
        if isinstance(v, (int, float)):
            return v
    return 0


# ------------------------------------------------------------ data-model metadata
def fetch_app_data_metadata(tenant, api_key, app_guid):
    """REST data-model metadata for ONE app: size, fields, tables.
    No engine / websocket session needed - reads the last-reload cached model.
    Returns the parsed dict (raises on HTTP error so the caller can record it)."""
    return _get_json(tenant, api_key, f"/api/v1/apps/{app_guid}/data/metadata", timeout=60)


def _summarize_metadata(md, with_field_detail=True):
    fields_raw = md.get("fields", []) or []
    tables_raw = md.get("tables", []) or []

    fields = [{
        "name": f.get("name", ""),
        "cardinality": _num(f, "cardinal", "cardinality", "distinct_value_count"),
        "total_count": _num(f, "total_count", "rows"),
        "byte_size": _num(f, "byte_size", "memory_size"),
        "is_system": bool(f.get("is_system", False)),
        "is_hidden": bool(f.get("is_hidden", False)),
        "tags": list(f.get("tags", []) or []),
        "src_tables": list(f.get("src_tables", []) or []),
    } for f in fields_raw]

    tables = [{
        "name": t.get("name", ""),
        "rows": _num(t, "no_of_rows", "rows"),
        "fields": _num(t, "no_of_fields", "fields"),
        "key_fields": _num(t, "no_of_key_fields"),
        "byte_size": _num(t, "byte_size", "memory_size"),
    } for t in tables_raw]

    out = {
        "size_bytes": _num(md, "static_byte_size", "byte_size"),
        "row_count": sum(t["rows"] for t in tables),
        "table_count": len(tables),
        "field_count": len(fields),
        "has_section_access": bool(md.get("has_section_access", False)),
        "is_direct_query": bool(md.get("is_direct_query_mode", False)),
    }
    if with_field_detail:
        out["fields"] = fields
        out["tables"] = tables
    return out


# ------------------------------------------------------------- data-file (storage)
def fetch_data_files_inventory(tenant, api_key, log=print, spaces=None):
    """ALL data files (QVDs etc.) across every space the key can see.

    Uses `includeAllSpaces=true&limit=1000` - the bare /data-files call defaults to
    the caller's PERSONAL space only (that's why earlier runs showed a handful of
    files). Each space has its own DataFiles connection; includeAllSpaces spans them.
    Resolves space + owner names. Note: Qlik can still omit a few very recently
    created files due to indexing - the tenant's own tooling has the same gap."""
    spaces = spaces or {}
    user_cache = {}
    files = []
    try:
        data = _get_json(tenant, api_key,
                         "/api/v1/data-files?includeAllSpaces=true&limit=1000", timeout=60)
    except Exception as e:
        log(f"  (data-files unavailable - storage view skipped: {e})")
        return files
    page = 0
    while True:
        for it in data.get("data", []):
            sid = it.get("spaceId", "") or ""
            files.append({
                "name": it.get("name", ""),
                "size_bytes": _num(it, "size", "byteSize"),
                "created": it.get("createdDate", "") or "",
                "modified": it.get("modifiedDate", "") or it.get("createdDate", "") or "",
                "space_id": sid,
                "space": spaces.get(sid, "Personal" if not sid else sid),
                "owner_id": it.get("ownerId", "") or "",
                "owner": _user_name(tenant, api_key, it.get("ownerId", ""), user_cache),
            })
        page += 1
        if page % 5 == 0:
            log(f"  ... {len(files)} data files listed")
        nxt = ((data.get("links", {}) or {}).get("next") or {}).get("href")
        if not nxt:
            break
        try:
            data = _get_url(api_key, nxt, timeout=60)
        except Exception:
            break
    log(f"Found {len(files)} data files across all visible spaces.")
    return files


# ------------------------------------------------------------------- main fetcher
def fetch_capacity_inventory(tenant, api_key, log=print,
                             with_field_detail=True, include_data_files=True,
                             max_apps=None, should_cancel=None):
    """Fetch a tenant-wide capacity & redundancy inventory.

    Returns:
      {
        "generated":  "YYYY-MM-DD HH:MM:SS",
        "tenant":     host,
        "totals":     {app_count, sized_app_count, total_app_bytes, data_file_bytes},
        "apps": [ {name, guid, space, space_id, owner, owner_id, last_reload,
                   created, modified, published, size_bytes, row_count,
                   table_count, field_count, has_section_access, is_direct_query,
                   fields:[...], tables:[...], error} ],
        "data_files": [ {name, size_bytes, created, modified, space_id, owner_id} ],
        "errors": [ {app, guid, error} ],
      }
    Set with_field_detail=False for a fast size-only sweep (skips per-field data).
    """
    host = normalize_host(tenant)
    log(f"Listing apps on {host} ...")
    apps = list_apps(tenant, api_key)
    try:
        spaces = list_spaces(tenant, api_key)
    except Exception as e:
        spaces = {}
        log(f"  (could not list spaces, names will be blank: {e})")
    if max_apps:
        apps = apps[:max_apps]
    log(f"Found {len(apps)} apps. Fetching size + field metadata (1-2 calls each) ...")

    user_cache = {}
    rows, errors = [], []
    for i, a in enumerate(apps, 1):
        _ck(should_cancel)
        guid, name = a["guid"], a.get("name", "")
        sid = a.get("space_id", "") or ""
        row = {
            "name": name, "guid": guid, "space_id": sid,
            "space": spaces.get(sid, "Personal" if not sid else sid),
            "owner": "", "owner_id": "", "last_reload": "", "created": "",
            "modified": "", "published": "", "size_bytes": 0, "row_count": 0,
            "table_count": 0, "field_count": 0, "has_section_access": False,
            "is_direct_query": False, "fields": [], "tables": [], "error": "",
        }
        # app attributes: reload time, dates, owner
        try:
            attrs = (_get_json(tenant, api_key, f"/api/v1/apps/{guid}")
                     .get("attributes", {}) or {})
            row["last_reload"] = attrs.get("lastReloadTime", "") or ""
            row["created"] = attrs.get("createdDate", "") or ""
            row["modified"] = attrs.get("modifiedDate", "") or ""
            row["published"] = attrs.get("publishTime", "") or ""
            row["owner_id"] = attrs.get("ownerId", "") or ""
            row["owner"] = _user_name(tenant, api_key, row["owner_id"], user_cache)
        except Exception as e:
            row["error"] = f"attributes: {getattr(e, 'reason', e)}"

        # data-model size + field/table cardinality
        try:
            md = fetch_app_data_metadata(tenant, api_key, guid)
            row.update(_summarize_metadata(md, with_field_detail))
        except Exception as e:
            msg = f"data/metadata: {getattr(e, 'reason', e)}"
            row["error"] = (row["error"] + "; " + msg) if row["error"] else msg
            errors.append({"app": name, "guid": guid, "error": msg})

        rows.append(row)
        if i % 25 == 0 or i == len(apps):
            log(f"  ... {i}/{len(apps)} apps")

    data_files = fetch_data_files_inventory(tenant, api_key, log) if include_data_files else []
    sized = [r for r in rows if r["size_bytes"]]
    totals = {
        "app_count": len(rows),
        "sized_app_count": len(sized),
        "total_app_bytes": sum(r["size_bytes"] for r in rows),
        "data_file_bytes": sum(f["size_bytes"] for f in data_files),
    }
    log(f"Done. {totals['sized_app_count']}/{totals['app_count']} apps sized, "
        f"{format_bytes(totals['total_app_bytes'])} of app data, "
        f"{len(data_files)} data files seen ({format_bytes(totals['data_file_bytes'])}).")
    return {
        "generated": datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "tenant": host, "totals": totals,
        "apps": rows, "data_files": data_files, "errors": errors,
    }


# --------------------------------------------------------------------- analysis
_COPY_RE = re.compile(r"^\s*(copy of |kopia av )+", re.I)
_SUFFIX_RE = re.compile(
    r"(\s*[-_ ]+(copy|kopia|backup|bak|old|gammal|test|dev|final|klar|v\d+|\(\d+\)))+\s*$",
    re.I)
_DATE_RE = re.compile(r"[\s_-]*\(?\d{4}[-_]\d{2}([-_]\d{2})?\)?\s*$")


def _norm_name(name):
    """Collapse 'Copy of X (2)', 'X_v3', 'X 2026-06' -> a shared base name so
    near-duplicate apps cluster together. Candidates only - confirm by eye."""
    s = (name or "").strip().lower()
    s = _COPY_RE.sub("", s)
    prev = None
    while prev != s:                  # peel repeated trailing copy/version/date tags
        prev = s
        s = _DATE_RE.sub("", s)
        s = _SUFFIX_RE.sub("", s)
    return re.sub(r"\s+", " ", s).strip()


# Only Personal space is genuinely free of Data-for-Analysis capacity. Apps in any
# Shared/Managed space - INCLUDING archive spaces - keep counting their last-reload
# data size until reloaded smaller or deleted, so they are NOT excluded; archive apps
# are flagged as prime reclaim candidates instead.
_DEFAULT_EXCLUDE_SPACES = ("Personal",)


def _space_excluded(space, excl_lower):
    return (space or "").strip().lower() in excl_lower


def _is_archive_space(space):
    """Archive-style spaces (Archive, Archived Apps, Archive - Finance, ...) - their
    apps are unused but STILL counting, so they are top reclaim candidates."""
    return (space or "").strip().lower().startswith("archive")


def _is_current_month(iso, ref=None):
    """True if the timestamp falls in the current calendar month (UTC)."""
    dt = _parse_dt(iso)
    if dt is None:
        return False
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=datetime.timezone.utc)
    now = ref or datetime.datetime.now(datetime.timezone.utc)
    return (dt.year, dt.month) == (now.year, now.month)


def analyze_capacity(inv, top_n=25, stale_days=120, shared_min_apps=3, recent_days=30,
                     exclude_spaces=_DEFAULT_EXCLUDE_SPACES, reloaded_this_month_only=False,
                     month_ref=None):
    """Turn a raw inventory into ranked savings / redundancy candidates:

      largest_apps             - biggest memory hogs (cut/trim these first)
      largest_fields           - heaviest single fields across all apps
      high_cardinality_fields  - distinct-value monsters (timestamps, unique IDs,
                                 free text) - prime split/drop candidates
      stale_large_apps         - big AND not reloaded in `stale_days` -> archive
      duplicate_app_clusters   - apps that normalise to the same base name
      shared_fields            - same field loaded in >= `shared_min_apps` apps
                                 (data paid for many times -> shared QVD layer)
    """
    excl = {s.strip().lower() for s in (exclude_spaces or ())}
    all_apps = [a for a in inv.get("apps", []) if a.get("size_bytes")]

    def keep(a):
        if _space_excluded(a.get("space"), excl):
            return False
        if reloaded_this_month_only and not _is_current_month(a.get("last_reload"), month_ref):
            return False
        return True

    apps = [a for a in all_apps if keep(a)]   # billable + reloaded this month
    space_excluded = [a for a in all_apps if _space_excluded(a.get("space"), excl)]
    month_excluded = [a for a in all_apps if not _space_excluded(a.get("space"), excl)
                      and reloaded_this_month_only
                      and not _is_current_month(a.get("last_reload"), month_ref)]

    largest_apps = sorted(apps, key=lambda a: -a["size_bytes"])[:top_n]

    flat = []
    for a in apps:
        for f in a.get("fields", []) or []:
            if f.get("is_system"):
                continue
            flat.append({**f, "app": a["name"], "guid": a["guid"], "space": a["space"]})
    largest_fields = sorted(flat, key=lambda f: -(f.get("byte_size") or 0))[:top_n]
    high_card = sorted(flat, key=lambda f: -(f.get("cardinality") or 0))[:top_n]

    # single-value fields (cardinality == 1): dead / placeholder columns - e.g. the
    # DW's 1901-01-01 null-placeholder on always-null date fields. Each is ~0 capacity
    # (one symbol, zero-bit pointers), but worth dropping at source for hygiene. Heavy
    # fields above are the capacity lever; this is correctness. Aggregated by name.
    sv = {}
    for f in flat:
        if (f.get("cardinality") or 0) == 1:
            d = sv.setdefault(f["name"].lower(),
                              {"field": f["name"], "app_count": 0, "spaces": set(),
                               "sample_app": "", "rows": 0, "date": False})
            d["app_count"] += 1
            d["spaces"].add(f.get("space", ""))
            if not d["sample_app"]:
                d["sample_app"] = f.get("app", "")
            d["rows"] = max(d["rows"], f.get("total_count") or 0)
            if any("date" in str(t).lower() or "timestamp" in str(t).lower()
                   for t in (f.get("tags") or [])):
                d["date"] = True
    single_value_fields = sorted(
        ({"field": d["field"], "app_count": d["app_count"],
          "spaces": ", ".join(list(sorted(s for s in d["spaces"] if s))[:5]),
          "sample_app": d["sample_app"], "rows": d["rows"], "date": d["date"]}
         for d in sv.values()), key=lambda x: -x["app_count"])[:200]

    now = datetime.datetime.now(datetime.timezone.utc)
    stale = []
    for a in apps:
        dt = _parse_dt(a.get("last_reload"))
        if dt is None:
            continue
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=datetime.timezone.utc)
        age = (now - dt).days
        if age >= stale_days:
            stale.append({"name": a["name"], "guid": a["guid"], "space": a["space"],
                          "owner": a["owner"], "size_bytes": a["size_bytes"],
                          "last_reload": a["last_reload"], "age_days": age})
    stale.sort(key=lambda a: -a["size_bytes"])

    clusters = {}
    for a in apps:
        clusters.setdefault(_norm_name(a["name"]), []).append(a)
    duplicate_clusters = []
    for base, group in clusters.items():
        if len(group) > 1:
            sizes = [g["size_bytes"] for g in group]
            total = sum(sizes)
            duplicate_clusters.append({
                "base_name": base, "count": len(group),
                "space_count": len({g["space"] for g in group}),
                "total_bytes": total,
                # what you'd reclaim by collapsing the copies down to ONE app
                "dedupe_savings_bytes": total - max(sizes),
                "newest_created": max((g.get("created", "") for g in group), default=""),
                "apps": [{"name": g["name"], "guid": g["guid"], "space": g["space"],
                          "owner": g.get("owner", ""), "size_bytes": g["size_bytes"],
                          "last_reload": g["last_reload"], "created": g.get("created", "")}
                         for g in sorted(group, key=lambda g: -g["size_bytes"])],
            })
    duplicate_clusters.sort(key=lambda c: -c["dedupe_savings_bytes"])

    # "What just changed?" view: apps created recently, flagged when they belong to
    # a duplicate set — a freshly rolled-out report copied across many spaces is the
    # classic capacity spike Sebastian's tenant is prone to.
    dup_bases = {c["base_name"] for c in duplicate_clusters}
    recent_large = []
    for a in apps:
        dt = _parse_dt(a.get("created"))
        if dt is None:
            continue
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=datetime.timezone.utc)
        age = (now - dt).days
        if 0 <= age <= recent_days:
            recent_large.append({"name": a["name"], "space": a["space"], "owner": a.get("owner", ""),
                                 "size_bytes": a["size_bytes"], "created": a.get("created", ""),
                                 "last_reload": a.get("last_reload", ""), "age_days": age,
                                 "in_duplicate_set": _norm_name(a["name"]) in dup_bases})
    recent_large.sort(key=lambda a: -a["size_bytes"])

    byfield = {}
    for f in flat:
        d = byfield.setdefault(f["name"].lower(),
                               {"name": f["name"], "apps": set(),
                                "total_bytes": 0, "max_card": 0})
        d["apps"].add(f["app"])
        d["total_bytes"] += f.get("byte_size") or 0
        d["max_card"] = max(d["max_card"], f.get("cardinality") or 0)
    shared_fields = sorted(
        ({"name": d["name"], "app_count": len(d["apps"]),
          "total_bytes": d["total_bytes"], "max_cardinality": d["max_card"],
          "apps": sorted(d["apps"])}
         for d in byfield.values() if len(d["apps"]) >= shared_min_apps),
        key=lambda d: -d["total_bytes"])

    # per-space rollup over ALL spaces (so the excluded Personal weight is visible),
    # each row flagged billable or not.
    space_roll = {}
    for a in all_apps:
        s = space_roll.setdefault(a["space"], {"space": a["space"], "app_count": 0, "bytes": 0,
                                               "billable": not _space_excluded(a.get("space"), excl)})
        s["app_count"] += 1
        s["bytes"] += a["size_bytes"]
    space_usage = sorted(space_roll.values(), key=lambda s: -s["bytes"])

    # consolidated ACTION LIST - the human-readable worklist of billable reclaim
    # candidates: unused (stale) apps + the redundant copies of duplicated reports.
    # Capacity CONTRIBUTORS worklist: every billable app that loads external data
    # OR creates an export (the only apps that touch capacity). Archive apps first
    # (delete candidates), then the rest by size. Notes carry archive/stale/duplicate
    # context. Apps that contribute ~0 (QVD-only / binary / no external & no export)
    # are skipped.
    dup_note = {}
    for c in duplicate_clusters:
        for g in c["apps"][1:]:                       # all but the largest copy
            if g.get("guid"):
                dup_note[g["guid"]] = f"duplicate copy ({c['count']}/{c['space_count']} spaces)"

    action = []
    for a in apps:
        scanned = a.get("scanned", False)
        ext = a.get("loads_external")
        exp = bool(a.get("creates_export"))
        archv = _is_archive_space(a.get("space"))
        age = _days_since(a.get("last_reload", ""))
        is_stale = isinstance(age, int) and age >= stale_days
        is_dup = a.get("guid") in dup_note
        counts = (ext is True) or exp                 # actually contributes to the bill
        confirmed_zero = scanned and (ext is False) and not exp
        reclaim = archv or is_stale or is_dup
        # keep real contributors, scanned-but-unread ('review'), and reclaim candidates;
        # drop small unscanned non-candidates and confirmed ~0 non-candidates.
        if not (counts or (scanned and ext is None) or reclaim):
            continue
        notes = []
        if confirmed_zero:
            notes.append("≈0 capacity (QVD/binary only)")
        if archv:
            notes.append("archive")
        if is_stale:
            notes.append(f"stale {age}d")
        if is_dup:
            notes.append(dup_note[a["guid"]])
        tier = 0 if counts else (2 if confirmed_zero else 1)
        action.append({"app": a["name"], "space": a["space"], "guid": a.get("guid", ""),
                       "owner": a.get("owner", ""), "size_bytes": a["size_bytes"],
                       "last_reload": a.get("last_reload", ""), "age_days": age,
                       "loads_external": ext, "source_kind": a.get("source_kind", ""),
                       "creates_export": exp, "is_archive": archv, "tier": tier,
                       "notes": "; ".join(notes)})
    # real contributors first, then 'review', then confirmed ~0 - biggest first in each
    action.sort(key=lambda x: (x["tier"], -(x["size_bytes"] or 0)))

    personal_summary = {
        "app_count": len(space_excluded),
        "bytes": sum(a["size_bytes"] for a in space_excluded),
        "spaces": sorted({a["space"] for a in space_excluded}),
        "month_excluded_count": len(month_excluded),
        "month_excluded_bytes": sum(a["size_bytes"] for a in month_excluded),
        "billable_count": len(apps),
        "billable_bytes": sum(a["size_bytes"] for a in apps),
    }

    return {
        "action_list": action,          # full contributor list (not capped)
        "largest_apps": largest_apps,
        "largest_fields": largest_fields,
        "high_cardinality_fields": high_card,
        "stale_large_apps": stale[:top_n],
        "duplicate_app_clusters": duplicate_clusters[:top_n],
        "recent_large_apps": recent_large[:top_n],
        "space_usage": space_usage,          # full list (one row per space)
        "personal_summary": personal_summary,
        "single_value_fields": single_value_fields,
        "shared_fields": shared_fields[:top_n],
    }


# ----------------------------------------------------------------- presentation
def format_bytes(n):
    n = float(n or 0)
    for unit in ("B", "KB", "MB", "GB", "TB"):
        if abs(n) < 1024 or unit == "TB":
            return f"{n:,.1f} {unit}" if unit != "B" else f"{int(n)} B"
        n /= 1024
    return f"{n:,.1f} TB"


def _days_since(iso):
    """Whole days between an ISO timestamp and now (UTC). '' if unparseable."""
    dt = _parse_dt(iso)
    if dt is None:
        return ""
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=datetime.timezone.utc)
    return (datetime.datetime.now(datetime.timezone.utc) - dt).days


def print_capacity_summary(inv, findings, top=10):
    """Quick console readout - the highest-value cut candidates first."""
    t = inv["totals"]
    print("\n=== QLIK CAPACITY INVENTORY ===")
    print(f"Tenant: {inv['tenant']}   Generated: {inv['generated']}")
    print(f"Apps sized: {t['sized_app_count']}/{t['app_count']}   "
          f"App data in memory: {format_bytes(t['total_app_bytes'])}   "
          f"Data files seen: {format_bytes(t['data_file_bytes'])}")

    print(f"\n-- Largest apps (top {top}) --")
    for a in findings["largest_apps"][:top]:
        print(f"  {format_bytes(a['size_bytes']):>11}  {a['name']}  "
              f"[{a['space']}]  reloaded {a['last_reload'][:10] or '-'}")

    print(f"\n-- Duplicate / near-duplicate app clusters (top {top}) --")
    for c in findings["duplicate_app_clusters"][:top]:
        print(f"  {format_bytes(c['total_bytes']):>11}  x{c['count']}  '{c['base_name']}'")
        for g in c["apps"]:
            print(f"               - {g['name']}  [{g['space']}]  "
                  f"{format_bytes(g['size_bytes'])}")

    print(f"\n-- Stale + large apps (no reload in a while) (top {top}) --")
    for a in findings["stale_large_apps"][:top]:
        print(f"  {format_bytes(a['size_bytes']):>11}  {a['age_days']:>4}d  "
              f"{a['name']}  [{a['space']}]")

    print(f"\n-- Heaviest fields across all apps (top {top}) --")
    for f in findings["largest_fields"][:top]:
        print(f"  {format_bytes(f['byte_size']):>11}  card={f['cardinality']:>10,}  "
              f"{f['name']}  ({f['app']})")

    print(f"\n-- Same field loaded in many apps (shared-QVD candidates) (top {top}) --")
    for f in findings["shared_fields"][:top]:
        print(f"  {format_bytes(f['total_bytes']):>11}  in {f['app_count']:>2} apps  "
              f"{f['name']}")


# ============================================================
#  IMPORT capacity - catalog datasets + raw data files
# ============================================================
def _basename_noext(name):
    return os.path.splitext(os.path.basename(str(name or "")))[0]


def _list_items(tenant, api_key, resource_type, log=print):
    """List catalog items of a given resourceType (e.g. 'dataset') via /items.
    Mirrors qlik_core.list_apps but for any resource type."""
    items, url = [], f"/api/v1/items?resourceType={resource_type}&limit=100"
    try:
        data = _get_json(tenant, api_key, url)
    except Exception as e:
        log(f"  (could not list {resource_type} items: {e})")
        return items
    page = 0
    while True:
        for it in data.get("data", []):
            rid = it.get("resourceId") or it.get("id")
            if rid:
                items.append({"id": rid, "name": it.get("name", "(unnamed)"),
                              "space_id": it.get("spaceId", "") or ""})
        page += 1
        if page % 20 == 0:        # heartbeat — a big catalog paginates for a while
            log(f"  ... listing {resource_type}s: {len(items)} so far")
        nxt = ((data.get("links", {}) or {}).get("next") or {}).get("href")
        if not nxt:
            break
        try:
            data = _get_url(api_key, nxt)
        except Exception:
            break
    return items


def fetch_dataset_inventory(tenant, api_key, log=print, spaces=None, max_detail=1000,
                            should_cancel=None):
    """Catalog datasets. Lists ALL dataset items (cheap, names + space — feeds
    tenant-wide duplicate detection), then reads `/data-sets/{id}` for size/rows on
    only the first `max_detail` items.

    Why the cap: the `/data-sets/{id}` endpoint is rate-limited (~100/min), so on a
    big catalog (tens of thousands of datasets) sizing every one would take hours and
    mostly hit 429s. `max_detail=None`/`0` sizes everything (use only on small
    tenants). The authoritative Import total comes from the QDI consumption meter, not
    from summing per-dataset sizes — this detail is for the redundancy worklist."""
    spaces = spaces or {}
    user_cache = {}
    items = _list_items(tenant, api_key, "dataset", log)
    total = len(items)
    detail_n = total if (max_detail is None or max_detail <= 0) else min(total, max_detail)
    if detail_n < total:
        log(f"Found {total} catalog datasets. Sizing the first {detail_n} only "
            f"(capped — the size endpoint is rate-limited). Duplicate detection still "
            f"covers all {total} by name.")
    else:
        log(f"Found {total} catalog datasets. Reading sizes ...")
    out = []
    for i, it in enumerate(items, 1):
        _ck(should_cancel)
        rec = {"name": it["name"], "id": it["id"], "space_id": it["space_id"],
               "space": spaces.get(it["space_id"], "Personal" if not it["space_id"] else it["space_id"]),
               "size_bytes": 0, "row_count": 0, "last_load": "", "type": "",
               "data_asset": "", "technical_name": "", "owner": "", "sized": False, "error": ""}
        if i <= detail_n:
            try:
                d = _get_json(tenant, api_key, f"/api/v1/data-sets/{it['id']}")
                op = d.get("operational", {}) or {}
                rec["size_bytes"] = _num(op, "size")
                rec["row_count"] = _num(op, "rowCount")
                rec["last_load"] = op.get("lastLoadTime", "") or ""
                rec["type"] = d.get("type", "") or ""
                rec["technical_name"] = d.get("technicalName", "") or ""
                rec["data_asset"] = (d.get("dataAssetInfo", {}) or {}).get("name", "") or ""
                rec["owner"] = _user_name(tenant, api_key, d.get("ownerId", ""), user_cache)
                rec["sized"] = True
            except Exception as e:
                rec["error"] = f"data-sets/{{id}}: {getattr(e, 'reason', e)}"
            if i % 25 == 0 or i == detail_n:
                log(f"  ... {i}/{detail_n} datasets sized")
        out.append(rec)
    return out


def fetch_import_inventory(tenant, api_key, log=print, include_data_files=True,
                           spaces=None, max_dataset_detail=1000, should_cancel=None):
    """Everything that consumes IMPORT capacity: catalog datasets + raw data files.
    Returns {datasets:[...], data_files:[...], totals:{...}}.
    Datasets and data files can overlap (a dataset backed by a QVD) - the totals
    are an upper bound; dedupe by name if you need an exact figure.
    `max_dataset_detail` caps per-dataset size reads (see fetch_dataset_inventory)."""
    if spaces is None:
        try:
            spaces = list_spaces(tenant, api_key)
        except Exception:
            spaces = {}
    datasets = fetch_dataset_inventory(tenant, api_key, log, spaces, max_detail=max_dataset_detail,
                                       should_cancel=should_cancel)
    files = fetch_data_files_inventory(tenant, api_key, log, spaces) if include_data_files else []
    return {
        "datasets": datasets, "data_files": files,
        "totals": {
            "dataset_count": len(datasets),
            "dataset_bytes": sum(d["size_bytes"] for d in datasets),
            "data_file_count": len(files),
            "data_file_bytes": sum(f["size_bytes"] for f in files),
        },
    }


def analyze_import_redundancy(import_inv, top_n=25, stale_days=120,
                              exclude_spaces=_DEFAULT_EXCLUDE_SPACES,
                              reloaded_this_month_only=True, month_ref=None):
    """Redundancy candidates on the import side:
      largest_datasets / largest_files  - heaviest single imports
      duplicate_datasets / duplicate_files - items that normalise to the same base
                                             name ('Copy of', '_v2', date stamps)
      stale_datasets    - not loaded in `stale_days` (archive candidates)
    Orphan detection (datasets/QVDs that NO app consumes) needs a lineage cross-ref
    - that's the natural next add once these land."""
    excl = {s.strip().lower() for s in (exclude_spaces or ())}
    # Stored QVDs / data files count toward capacity for as long as they exist,
    # regardless of when last loaded - so the current-month rule (which is about APP
    # reloads) is NOT applied here; only the space exclusion is.
    ds = [d for d in import_inv.get("datasets", [])
          if d.get("size_bytes") and not _space_excluded(d.get("space"), excl)]
    files = [f for f in import_inv.get("data_files", [])
             if f.get("size_bytes") and not _space_excluded(f.get("space"), excl)]

    def dup_clusters(rows):
        cl = {}
        for r in rows:
            cl.setdefault(_norm_name(_basename_noext(r["name"])), []).append(r)
        out = [{"base_name": k, "count": len(g),
                "total_bytes": sum(x["size_bytes"] for x in g),
                "items": sorted(g, key=lambda x: -x["size_bytes"])}
               for k, g in cl.items() if len(g) > 1]
        out.sort(key=lambda c: -c["total_bytes"])
        return out[:top_n]

    now = datetime.datetime.now(datetime.timezone.utc)
    stale = []
    for d in ds:
        dt = _parse_dt(d.get("last_load"))
        if dt is None:
            continue
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=datetime.timezone.utc)
        age = (now - dt).days
        if age >= stale_days:
            stale.append({"name": d["name"], "space": d["space"],
                          "size_bytes": d["size_bytes"], "last_load": d["last_load"],
                          "age_days": age})
    stale.sort(key=lambda d: -d["size_bytes"])

    return {
        "largest_datasets": sorted(ds, key=lambda d: -d["size_bytes"])[:top_n],
        "largest_files": sorted(files, key=lambda f: -f["size_bytes"])[:top_n],
        "duplicate_datasets": dup_clusters(ds),
        "duplicate_files": dup_clusters(files),
        "stale_datasets": stale[:top_n],
    }


# ============================================================
#  Authoritative consumption (tenant-admin) - the real meter
# ============================================================
def fetch_consumption(tenant, api_key, periods=("current", "previous"), log=print):
    """Entitlement-consumption records from /api/v1/consumption/executions
    (needs tenant-admin / consumption access). Returns the raw record list.
    Raises on HTTP error (e.g. 403 if the key lacks the entitlement) so the
    caller can fall back to the proxy inventories."""
    q = "limit=200" + "".join(f"&periodsToInclude={p}" for p in periods)
    recs = []
    data = _get_json(tenant, api_key, f"/api/v1/consumption/executions?{q}", timeout=60)
    while True:
        recs.extend(data.get("data", []) or [])
        nxt = ((data.get("links", {}) or {}).get("next") or {}).get("href")
        if not nxt:
            break
        try:
            data = _get_url(api_key, nxt, timeout=60)
        except Exception:
            break
    return recs


def summarize_consumption(records):
    """Flatten consumption records and surface the two capacities. Each record's
    `segments` (e.g. {'APP': n} for app reloads, {'QDI': n} for data-integration
    imports) carries the Data-for-Analysis split; the record also carries
    capacityLimit / localUsage / overage per period.

    NOTE: confirm the field mapping on your first run - Qlik's exact record shape
    can vary by subscription. The full parsed period list is returned for review."""
    periods = []
    for r in records:
        seg = {}
        for s in (r.get("segments", []) or []):
            if isinstance(s, dict):
                seg.update(s)
        periods.append({
            "periodType": r.get("periodType", ""), "periodStart": r.get("periodStart", ""),
            "periodEnd": r.get("periodEnd", ""), "resourceType": r.get("resourceType", ""),
            "resourceAction": r.get("resourceAction", ""), "taskName": r.get("taskName", ""),
            "unit": r.get("unit", ""), "capacityLimit": r.get("capacityLimit"),
            "localUsage": r.get("localUsage"), "globalUsage": r.get("globalUsage"),
            "overage": bool(r.get("overage")), "closeToOverage": bool(r.get("closeToOverage")),
            "segments": seg,
        })

    def latest_with_segment(key):
        cand = [p for p in periods if key in p["segments"]]
        cand.sort(key=lambda p: (p["periodType"] == "month", p.get("periodEnd", "")), reverse=True)
        return cand[0] if cand else None

    def find_billed_meter():
        # The actual Data-for-Analysis meter on this tenant. On Bufab it surfaces as
        # resourceType 'data.volume.consumption' / task 'dataVolumeAggregated', with
        # usage + limit in BYTES — NOT as APP/QDI segments. Pick the record with the
        # largest capacityLimit among those as the headline.
        cand = [p for p in periods
                if "datavolume" in (p.get("taskName", "") or "").lower()
                or "data.volume" in (p.get("resourceType", "") or "").lower()]
        cand.sort(key=lambda p: (p.get("capacityLimit") or 0, p.get("periodEnd", "")), reverse=True)
        return cand[0] if cand else None

    return {"data_volume": find_billed_meter(),
            "app_reload": latest_with_segment("APP"),
            "import_qdi": latest_with_segment("QDI"),
            "periods": periods}


# ============================================================
#  Orphan detection - imports that NO app consumes
# ============================================================
def build_consumption_index(tenant, api_key, apps=None, log=print, max_apps=None,
                            should_cancel=None):
    """Open every app, read its load script, and record which data files it
    READS vs PRODUCES. This is what lets us spot imports nothing consumes.

    Returns {consumed: set(basenames read by >=1 app),
             produced: set(basenames written by a STORE),
             scripts_read: int, errors: [{app, guid, error}]}.

    Opens each app in the engine to call GetScript - the same approach the tool's
    cross-app index already uses. For a big tenant this is the slow step. Coverage
    equals the scripts successfully read; apps that error are listed so a file they
    read isn't wrongly called an orphan."""
    if apps is None:
        apps = list_apps(tenant, api_key)
    if max_apps:
        apps = apps[:max_apps]
    consumed, produced, errors = set(), set(), []
    log(f"Reading load scripts of {len(apps)} apps for the consumption index ...")
    for i, a in enumerate(apps, 1):
        _ck(should_cancel)
        guid = a["guid"]
        exp = QlikExporter(tenant, api_key, guid, ".", log=lambda *_: None)
        script = ""
        try:
            exp.connect()
            app_h = exp.call(-1, "OpenDoc", [guid])["qReturn"]["qHandle"]
            script = exp.fetch_script(app_h)
        except Exception as e:
            errors.append({"app": a.get("name", ""), "guid": guid,
                           "error": str(getattr(e, "reason", e))[:200]})
        finally:
            exp.close()
        if script:
            stores, _reads = parse_store_reads(script)          # STORE ... INTO *.qvd
            produced |= stores
            consumed |= (extract_file_refs(script) - stores)    # everything read, not written
        if i % 25 == 0 or i == len(apps):
            log(f"  ... {i}/{len(apps)} scripts")
    return {"consumed": consumed, "produced": produced,
            "scripts_read": len(apps) - len(errors), "errors": errors}


def detect_orphans(import_inv, index, exclude_spaces=("Personal",)):
    """Classify each imported artifact against the consumption index:
      data files: consumed | produced_only | orphan ; datasets: consumed | orphan_candidate.
    Files/datasets in excluded spaces (Personal) are SKIPPED - they do not count toward
    capacity. The result carries `coverage` + `low_confidence`: when the load-script
    scan covered only part of the apps, a QVD read solely by an unread app is wrongly
    flagged orphan, so trust the lists only when low_confidence is False."""
    excl = {s.strip().lower() for s in (exclude_spaces or ())}
    consumed = index.get("consumed", set())
    produced = index.get("produced", set())

    file_rows = []
    for f in import_inv.get("data_files", []):
        if _space_excluded(f.get("space"), excl):
            continue
        base = os.path.basename(f.get("name", "")).lower()
        status = "consumed" if base in consumed else \
                 ("produced_only" if base in produced else "orphan")
        file_rows.append({**f, "basename": base, "status": status})

    def _names(d):
        out = set()
        for v in (d.get("name", ""), d.get("technical_name", "")):
            if v:
                b = os.path.basename(str(v)).lower()
                out.add(b)
                out.add(os.path.splitext(b)[0])
        return out

    ds_rows = []
    for d in import_inv.get("datasets", []):
        if _space_excluded(d.get("space"), excl):
            continue
        names = _names(d)
        hit = any(n in consumed or (n + ".qvd") in consumed for n in names)
        ds_rows.append({**d, "status": "consumed" if hit else "orphan_candidate"})

    orphan_files = sorted((r for r in file_rows if r["status"] == "orphan"),
                          key=lambda r: -(r.get("size_bytes") or 0))
    produced_only = sorted((r for r in file_rows if r["status"] == "produced_only"),
                           key=lambda r: -(r.get("size_bytes") or 0))
    orphan_ds = sorted((r for r in ds_rows if r["status"] == "orphan_candidate"),
                       key=lambda r: -(r.get("size_bytes") or 0))

    return {
        "orphan_files": orphan_files,
        "produced_only_files": produced_only,
        "orphan_datasets": orphan_ds,
        "reclaimable": {
            "orphan_file_bytes": sum(r.get("size_bytes") or 0 for r in orphan_files),
            "produced_only_bytes": sum(r.get("size_bytes") or 0 for r in produced_only),
            "orphan_dataset_bytes": sum(r.get("size_bytes") or 0 for r in orphan_ds),
        },
        "index_scripts_read": index.get("scripts_read", 0),
        "index_errors": index.get("errors", []),
        "coverage": (index.get("scripts_read", 0) / index["scripts_total"]
                     if index.get("scripts_total") else 1.0),
        "low_confidence": bool(index.get("scripts_total")
                               and index.get("scripts_read", 0) / index["scripts_total"] < 0.97),
    }


# ============================================================
#  External-load detection - does an app actually count?
# ============================================================
def _classify_external_load(script):
    """(loads_external, source_kind) from a load script (best-effort text parse).

    Data for Analysis counts the EXTERNAL data an app ingests. An app that only reads
    QVDs/files already in Qlik, or binary-loads another app, adds ~0 (its sources are
    counted elsewhere). A database/SQL source is the clear 'counts' signal.
      True  -> loads external data (counts toward capacity)
      False -> QVD/file-only or binary load (~0)
      None  -> file-based, can't tell from text (review)"""
    s = script or ""
    tables = parse_load_tables(s)
    if any(t.get("kind") == "sql" for t in tables):
        return True, "external DB (SQL)"
    if re.search(r"(?im)^\s*binary\b", s):
        return False, "binary load"
    froms = [t for t in tables if t.get("kind") == "from"]
    if not froms:
        return False, "no external load"
    if all(all(str(f).endswith(".qvd") for f in (t.get("files") or [])) for t in froms):
        return False, "QVD/file only"
    return None, "file-based (review)"


def scan_app_load_profiles(tenant, api_key, apps, log=print, max_apps=None, should_cancel=None):
    """Read each app's load script ONCE and return its capacity profile:
      {guid: {loads_external, source_kind, creates_export, stores:set, reads:set}}
    - loads_external -> the app contributes to the APP bucket (external data ingest)
    - creates_export -> the QVDs it STOREs contribute to the datafile bucket
    - stores/reads also feed orphan detection, so this single pass serves both.
    Slow: one engine GetScript per app. With max_apps=None it reads every app."""
    items = [a for a in apps if a.get("guid")]
    if max_apps:
        items = items[:max_apps]
    out, errors = {}, 0
    log(f"Reading {len(items)} app load scripts (external-load + export profile) ...")
    for i, a in enumerate(items, 1):
        _ck(should_cancel)
        guid = a["guid"]
        script = ""
        for attempt in range(2):                 # one retry for a transient blip
            exp = QlikExporter(tenant, api_key, guid, ".", log=lambda *_: None)
            try:
                exp.connect()
                app_h = exp.call(-1, "OpenDoc", [guid])["qReturn"]["qHandle"]
                script = exp.fetch_script(app_h)
                break
            except Exception:
                if attempt == 0:
                    time.sleep(1.0)
            finally:
                exp.close()
        if not script:
            errors += 1
        if script:
            ext, kind = _classify_external_load(script)
            stores, _r = parse_store_reads(script)
            reads = extract_file_refs(script) - stores
        else:
            ext, kind, stores, reads = None, "unread", set(), set()
        out[guid] = {"loads_external": ext, "source_kind": kind,
                     "creates_export": bool(stores), "stores": stores, "reads": reads}
        if i % 50 == 0 or i == len(items):
            log(f"  ... {i}/{len(items)} scripts")
    if errors:
        log(f"  ({errors} scripts could not be read - those apps show as 'review')")
    return out


# ============================================================
#  Orchestrator - both capacities in one call
# ============================================================
def fetch_two_capacities(tenant, api_key, log=print, with_field_detail=True,
                         with_orphans=False, max_dataset_detail=1000,
                         max_orphan_apps=400, exclude_spaces=_DEFAULT_EXCLUDE_SPACES,
                         reloaded_this_month_only=False, with_external_detection=True,
                         max_external_scan=600, should_cancel=None):
    """Gather BOTH capacities in one pass:
      App reload : per-app data-model inventory (proxy) + consumption (authoritative)
      Import     : catalog datasets + data files (proxy) + consumption (authoritative)

    Scale guardrails (this is a tenant-wide sweep — big tenants have tens of
    thousands of datasets and thousands of apps):
      - `max_dataset_detail` caps per-dataset size reads (rate-limited endpoint);
        duplicate detection still covers every dataset by name.
      - `with_orphans` reads EVERY app's load script over the engine; it is skipped
        when the app count exceeds `max_orphan_apps`, because partial coverage would
        mislabel files as orphans. Re-run on a subset to use it on a big tenant.
    Consumption is best-effort: without a tenant-admin key it is skipped and you
    still get the proxy ranking. For a quick "which meter is over?" use scan_meters()."""
    try:
        spaces = list_spaces(tenant, api_key)
    except Exception:
        spaces = {}

    log("== APP RELOAD capacity: app data-model inventory ==")
    app_inv = fetch_capacity_inventory(tenant, api_key, log=log,
                                       with_field_detail=with_field_detail,
                                       include_data_files=False, should_cancel=should_cancel)
    profiles = {}
    if with_external_detection:
        scope = "all apps" if not max_external_scan else f"top {max_external_scan} by size"
        log(f"== LOAD-SCRIPT scan: external-load + export ({scope}) ==")
        targets = sorted(app_inv["apps"], key=lambda a: -(a.get("size_bytes") or 0))
        profiles = scan_app_load_profiles(tenant, api_key, targets, log=log,
                                          max_apps=max_external_scan, should_cancel=should_cancel)
        for a in app_inv["apps"]:
            p = profiles.get(a.get("guid"))
            if p:
                a["scanned"] = True
                a["loads_external"] = p["loads_external"]
                a["source_kind"] = p["source_kind"]
                a["creates_export"] = p["creates_export"]
    ar_red = analyze_capacity(app_inv, exclude_spaces=exclude_spaces,
                              reloaded_this_month_only=reloaded_this_month_only)

    _ck(should_cancel)
    log("== IMPORT capacity: catalog datasets + data files ==")
    imp_inv = fetch_import_inventory(tenant, api_key, log=log, spaces=spaces,
                                     max_dataset_detail=max_dataset_detail,
                                     should_cancel=should_cancel)
    imp_section = {"inventory": imp_inv,
                   "redundancy": analyze_import_redundancy(
                       imp_inv, exclude_spaces=exclude_spaces,
                       reloaded_this_month_only=reloaded_this_month_only)}

    if with_orphans:
        if profiles:
            # reuse the load-script scan already done - no second crawl
            consumed, produced = set(), set()
            for p in profiles.values():
                produced |= p.get("stores", set())
                consumed |= p.get("reads", set())
            total = len([a for a in app_inv["apps"] if a.get("guid")])
            read_ok = sum(1 for p in profiles.values() if p.get("source_kind") != "unread")
            coverage = (read_ok / total) if total else 0.0
            if coverage >= 0.90:
                index = {"consumed": consumed, "produced": produced,
                         "scripts_read": read_ok, "scripts_total": total, "errors": []}
                imp_section["orphans"] = detect_orphans(imp_inv, index, exclude_spaces=exclude_spaces)
            else:
                log(f"== ORPHAN detection SKIPPED: only {coverage:.0%} of app scripts read "
                    f"({read_ok}/{total}). Partial coverage flags nearly everything as 'orphan', "
                    "so it is omitted. Re-run with max_external_scan=None on a stable connection "
                    "to get reliable orphans. ==")
        else:
            n_apps = len(app_inv["apps"])
            if n_apps > max_orphan_apps:
                log(f"== ORPHAN detection SKIPPED: {n_apps} apps exceeds the cap "
                    f"({max_orphan_apps}); turn on external detection or raise the cap. ==")
            else:
                log("== ORPHAN detection: reading app load scripts ==")
                index = build_consumption_index(tenant, api_key, apps=app_inv["apps"], log=log,
                                                should_cancel=should_cancel)
                index["scripts_total"] = len(app_inv["apps"])
                imp_section["orphans"] = detect_orphans(imp_inv, index, exclude_spaces=exclude_spaces)

    _ck(should_cancel)
    log("== Authoritative consumption (tenant-admin) ==")
    consumption = None
    try:
        consumption = summarize_consumption(fetch_consumption(tenant, api_key, log=log))
    except Exception as e:
        log(f"  (consumption endpoint unavailable - proxies only: {getattr(e, 'reason', e)})")

    return {
        "generated": app_inv["generated"], "tenant": app_inv["tenant"],
        "app_reload": {"inventory": app_inv, "redundancy": ar_red},
        "import": imp_section,
        "consumption": consumption,
    }


def print_two_capacity_summary(result, top=10):
    c = result.get("consumption") or {}
    print("\n================ QLIK CAPACITY - two meters ================")
    print(f"Tenant: {result['tenant']}   Generated: {result['generated']}")

    def show_meter(title, rec):
        if not rec:
            print(f"\n[{title}] no authoritative record returned "
                  "(admin endpoint empty or not entitled).")
            return
        lim = rec.get("capacityLimit")
        val = next(iter(rec.get("segments", {}).values()), rec.get("localUsage"))
        pct = f"{val / lim * 100:.0f}%" if (isinstance(val, (int, float)) and lim) else "?"
        flag = "  !! OVERAGE" if rec.get("overage") else \
               ("  ! close to limit" if rec.get("closeToOverage") else "")
        print(f"\n[{title}] {rec.get('periodStart', '')[:10]}..{rec.get('periodEnd', '')[:10]}  "
              f"usage {val} / limit {lim} ({pct}) {rec.get('unit', '')}{flag}")

    dv = c.get("data_volume")
    if dv:
        used, lim = dv.get("localUsage"), dv.get("capacityLimit")
        pct = f"{used / lim * 100:.3f}%" if (isinstance(used, (int, float)) and lim) else "?"
        over = (used - lim) if (isinstance(used, (int, float)) and isinstance(lim, (int, float))) else None
        flag = (f"  !! OVERAGE (over by {format_bytes(over)})" if dv.get("overage") and over and over > 0
                else ("  ! close to limit" if dv.get("closeToOverage") else ""))
        print(f"\n[BILLED Data for Analysis] {format_bytes(used)} / {format_bytes(lim)} ({pct}){flag}")
    else:
        print("\n[BILLED Data for Analysis] no dataVolume record found")
    if c.get("app_reload"):
        show_meter("APP segment", c.get("app_reload"))
    if c.get("import_qdi"):
        show_meter("QDI segment", c.get("import_qdi"))

    ar = result["app_reload"]; ai, arr = ar["inventory"], ar["redundancy"]
    print(f"\n--- APP RELOAD redundancy "
          f"(proxy: {format_bytes(ai['totals']['total_app_bytes'])} across "
          f"{ai['totals']['sized_app_count']} apps) ---")
    print(" Largest apps:")
    for a in arr["largest_apps"][:top]:
        print(f"   {format_bytes(a['size_bytes']):>11}  {a['name']}  [{a['space']}]")
    print(" Duplicate app clusters (same report across spaces):")
    for cl in arr["duplicate_app_clusters"][:top]:
        print(f"   reclaim {format_bytes(cl.get('dedupe_savings_bytes')):>11}  "
              f"x{cl['count']} copies / {cl.get('space_count')} spaces  '{cl['base_name']}'")
    print(" Top spaces by app data:")
    for s in arr.get("space_usage", [])[:top]:
        print(f"   {format_bytes(s['bytes']):>11}  {s['app_count']:>4} apps  {s['space']}")

    im = result["import"]; ii, irr = im["inventory"], im["redundancy"]
    print(f"\n--- IMPORT redundancy "
          f"(proxy: {format_bytes(ii['totals']['dataset_bytes'])} datasets + "
          f"{format_bytes(ii['totals']['data_file_bytes'])} files) ---")
    print(" Largest datasets:")
    for d in irr["largest_datasets"][:top]:
        print(f"   {format_bytes(d['size_bytes']):>11}  {d['name']}  [{d['space']}]")
    print(" Duplicate datasets:")
    for cl in irr["duplicate_datasets"][:top]:
        print(f"   {format_bytes(cl['total_bytes']):>11}  x{cl['count']}  '{cl['base_name']}'")
    print(" Duplicate / superseded files:")
    for cl in irr["duplicate_files"][:top]:
        print(f"   {format_bytes(cl['total_bytes']):>11}  x{cl['count']}  '{cl['base_name']}'")

    orph = im.get("orphans")
    if orph:
        rc = orph["reclaimable"]
        print(f"\n--- IMPORT ORPHANS (cross-referenced {orph['index_scripts_read']} app scripts) ---")
        print(f" Orphan files - no app reads or writes them: "
              f"{format_bytes(rc['orphan_file_bytes'])} across {len(orph['orphan_files'])}")
        for r in orph["orphan_files"][:top]:
            print(f"   {format_bytes(r['size_bytes']):>11}  {r['name']}")
        print(f" Produced-but-unread - written by a reload, read by nothing: "
              f"{format_bytes(rc['produced_only_bytes'])} across {len(orph['produced_only_files'])}")
        for r in orph["produced_only_files"][:top]:
            print(f"   {format_bytes(r['size_bytes']):>11}  {r['name']}")
        print(f" Orphan dataset candidates (verify - refs can be indirect): "
              f"{format_bytes(rc['orphan_dataset_bytes'])} across {len(orph['orphan_datasets'])}")
        for r in orph["orphan_datasets"][:top]:
            print(f"   {format_bytes(r['size_bytes']):>11}  {r['name']}  [{r['space']}]")
        if orph["index_errors"]:
            print(f"   (note: {len(orph['index_errors'])} app scripts could not be read - "
                  "their reads aren't counted, so verify those before deleting)")


# ============================================================
#  Access / coverage - which spaces errored (no access vs no data)
# ============================================================
def summarize_errors_by_space(app_inv):
    """Group app errors by space so access gaps are visible. Distinguishes
    ACCESS-DENIED (403/401 - you can't read that space) from NO-METADATA
    (404 / never reloaded - not an access problem). Returns rows sorted with the
    access-denied spaces first."""
    roll = {}
    for a in app_inv.get("apps", []):
        err = (a.get("error") or "").strip()
        if not err:
            continue
        space = a.get("space") or "(unknown)"
        s = roll.setdefault(space, {"space": space, "errored": 0, "access_denied": 0,
                                    "no_metadata": 0, "other": 0, "sample": ""})
        s["errored"] += 1
        el = err.lower()
        if "forbidden" in el or "unauthorized" in el or "403" in el or "401" in el:
            s["access_denied"] += 1
        elif "not found" in el or "404" in el:
            s["no_metadata"] += 1
        else:
            s["other"] += 1
        if not s["sample"]:
            s["sample"] = err[:200]
    return sorted(roll.values(), key=lambda x: (-x["access_denied"], -x["errored"]))


# ============================================================
#  Excel report - one workbook, a sheet per view
# ============================================================
def write_capacity_report(result, out_dir, log=print):
    """Write a multi-sheet Excel workbook from a fetch_two_capacities() result,
    matching the style of write_usage_report (header fills, freeze panes, auto-
    filter, '='-string coercion). Falls back to one CSV per sheet if openpyxl is
    missing. Returns the main output path.

    Sheets: Summary (authoritative meters + proxy totals + reclaimable), then one
    sheet each for the App-reload and Import redundancy / orphan views."""
    os.makedirs(out_dir, exist_ok=True)
    stamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    base = f"capacity_report_{stamp}"

    ar = result.get("app_reload", {}) or {}
    im = result.get("import", {}) or {}
    cons = result.get("consumption") or {}
    ai, arr = ar.get("inventory", {}) or {}, ar.get("redundancy", {}) or {}
    ii, irr = im.get("inventory", {}) or {}, im.get("redundancy", {}) or {}
    orph = im.get("orphans") or {}
    tot_app = ai.get("totals", {}).get("total_app_bytes") or 0
    space_errors = summarize_errors_by_space(ai)

    def mb(n):
        return round((n or 0) / 1024 / 1024, 1)

    def pct_app(n):
        return round((n or 0) / tot_app * 100, 2) if tot_app else ""

    def meter_rows(lbl, rec):
        if not rec:
            return [[f"{lbl} - authoritative", "no record (admin endpoint empty or not entitled)"]]
        seg = rec.get("segments", {}) or {}
        val = next(iter(seg.values()), rec.get("localUsage"))
        status = "OVERAGE" if rec.get("overage") else \
                 ("close to limit" if rec.get("closeToOverage") else "ok")
        return [
            [f"{lbl} - period", f"{str(rec.get('periodStart', ''))[:10]} .. "
                                f"{str(rec.get('periodEnd', ''))[:10]}"],
            [f"{lbl} - usage", val],
            [f"{lbl} - capacity limit", rec.get("capacityLimit")],
            [f"{lbl} - status", status],
        ]

    dv = cons.get("data_volume")
    persum = arr.get("personal_summary", {}) or {}
    total_app = ai.get("totals", {}).get("total_app_bytes") or 0
    personal_bytes = persum.get("bytes", 0) or 0

    incomplete = (dv is None) and not (ii.get("totals", {}).get("dataset_count")
                                       or ii.get("totals", {}).get("data_file_count"))
    unread = sum(1 for a in arr.get("action_list", []) if a.get("loads_external") is None)

    summary_rows = [
        ["Generated", result.get("generated", stamp)],
        ["Tenant", result.get("tenant", "")],
    ]
    if incomplete:
        summary_rows.append(
            ["!! RUN INCOMPLETE", "network errors (billed meter / datasets / files "
             "unavailable) - re-run on a stable connection"])
    summary_rows += [
        ["", ""],
        ["== BILLED CAPACITY (Data for Analysis) ==", ""],
    ]
    if dv:
        used, lim = dv.get("localUsage"), dv.get("capacityLimit")
        over = (used - lim) if (isinstance(used, (int, float)) and isinstance(lim, (int, float))) else None
        status = (f"OVERAGE - over by {format_bytes(over)}" if dv.get("overage") and over and over > 0
                  else ("close to limit" if dv.get("closeToOverage") else "ok"))
        summary_rows += [
            ["Period", f"{str(dv.get('periodStart', ''))[:10]} .. {str(dv.get('periodEnd', ''))[:10]}"],
            ["Usage", format_bytes(used)],
            ["Capacity limit", format_bytes(lim)],
            ["Used %", (f"{used / lim * 100:.3f}%" if lim else "")],
            ["Status", status],
        ]
    else:
        summary_rows += [["Billed meter", "no dataVolume record (need tenant-admin / see Consumption raw)"]]
    for seglbl, rec in (("App reload (APP segment)", cons.get("app_reload")),
                        ("Import/QDI (QDI segment)", cons.get("import_qdi"))):
        if rec:
            summary_rows += meter_rows(seglbl, rec)
    summary_rows += [
        ["", ""],
        ["== APP-SIDE PROXY (in-memory size; NOT the billed metric) ==", ""],
        [f"Billable apps ({persum.get('billable_count', 0)}) (MB)",
         mb(persum.get('billable_bytes', 0))],
        [f"Excluded - Personal only ({persum.get('app_count', 0)}) (MB)",
         mb(persum.get('bytes', 0))],
        ["Apps sized (all)", ai.get("totals", {}).get("sized_app_count")],
        ["Duplicate-report reclaim if consolidated (MB)",
         mb(sum(c.get("dedupe_savings_bytes", 0) for c in arr.get("duplicate_app_clusters", [])))],
        ["", ""],
        ["== IMPORT-SIDE PROXY ==", ""],
        ["Catalog datasets sized, all spaces (MB)", mb(ii.get("totals", {}).get("dataset_bytes"))],
        ["Datasets (listed)", ii.get("totals", {}).get("dataset_count")],
        ["Data files (MB)", mb(ii.get("totals", {}).get("data_file_bytes"))],
        ["Data files", ii.get("totals", {}).get("data_file_count")],
        ["", ""],
        ["== ACCESS / COVERAGE ==", ""],
        ["Spaces with ACCESS-DENIED errors (no access)",
         sum(1 for e in space_errors if e["access_denied"])],
        ["Spaces with any app errors", len(space_errors)],
        ["Apps that errored (unsized)", sum(e["errored"] for e in space_errors)],
        ["Apps unread - external scan failed ('review')", unread],
    ]
    if orph:
        rc = orph.get("reclaimable", {})
        summary_rows += [
            ["", ""],
            ["== RECLAIMABLE (orphans, Personal excluded) ==", ""],
            ["Orphan files (MB)", mb(rc.get("orphan_file_bytes"))],
            ["Produced-but-unread (MB)", mb(rc.get("produced_only_bytes"))],
            ["Orphan dataset candidates (MB)", mb(rc.get("orphan_dataset_bytes"))],
            ["App scripts cross-referenced", orph.get("index_scripts_read")],
            ["Coverage", f"{orph.get('coverage', 1.0) * 100:.0f}%"],
        ]
        if orph.get("low_confidence"):
            summary_rows.append(["!! LOW CONFIDENCE",
                                 "partial script coverage - some 'orphans' may be read by "
                                 "unscanned apps; verify before deleting"])

    sheets = [
        ("Action list (billable)",
         ["App", "Space", "Owner", "Size (MB)", "Last reload", "Age (days)",
          "Loads external?", "Source", "Creates QVD?", "Notes"],
         [[a["app"], a.get("space", ""), a.get("owner", ""), mb(a.get("size_bytes")),
           str(a.get("last_reload", ""))[:19], a.get("age_days"),
           ("yes" if a.get("loads_external") is True else
            ("no" if a.get("loads_external") is False else "review")),
           a.get("source_kind", ""),
           ("yes" if a.get("creates_export") else "no"),
           a.get("notes", "")]
          for a in arr.get("action_list", [])]),
        ("App reload - largest apps",
         ["App", "Space", "Owner", "Size (MB)", "Rows", "Tables", "Fields",
          "Last reload", "Days since reload"],
         [[a["name"], a.get("space", ""), a.get("owner", ""), mb(a.get("size_bytes")),
           a.get("row_count"), a.get("table_count"), a.get("field_count"),
           str(a.get("last_reload", ""))[:19], _days_since(a.get("last_reload", ""))]
          for a in arr.get("largest_apps", [])]),
        ("App reload - by space",
         ["Space", "Apps", "Size (MB)", "% of app data"],
         [[s["space"], s["app_count"], mb(s["bytes"]), pct_app(s["bytes"])]
          for s in arr.get("space_usage", [])]),
        ("App reload - duplicates",
         ["Base name", "Copies", "Spaces", "Reclaim if 1 kept (MB)", "App", "Space", "Owner",
          "Size (MB)", "Created", "Last reload", "Days since reload"],
         [[c["base_name"], c.get("count"), c.get("space_count"),
           mb(c.get("dedupe_savings_bytes", 0)), g["name"], g.get("space", ""), g.get("owner", ""),
           mb(g.get("size_bytes")), str(g.get("created", ""))[:10],
           str(g.get("last_reload", ""))[:19], _days_since(g.get("last_reload", ""))]
          for c in arr.get("duplicate_app_clusters", []) for g in c["apps"]]),
        ("App reload - recent large adds",
         ["App", "Space", "Owner", "Size (MB)", "Created", "Last reload", "Age (days)",
          "In duplicate set?"],
         [[a["name"], a.get("space", ""), a.get("owner", ""), mb(a.get("size_bytes")),
           str(a.get("created", ""))[:10], str(a.get("last_reload", ""))[:19], a.get("age_days"),
           "Yes" if a.get("in_duplicate_set") else ""]
          for a in arr.get("recent_large_apps", [])]),
        ("App reload - stale large",
         ["App", "Space", "Owner", "Size (MB)", "Age (days)", "Last reload"],
         [[a["name"], a.get("space", ""), a.get("owner", ""), mb(a.get("size_bytes")),
           a.get("age_days"), str(a.get("last_reload", ""))[:19]]
          for a in arr.get("stale_large_apps", [])]),
        ("App reload - heavy fields",
         ["Field", "App", "Space", "Byte size (MB)", "Cardinality"],
         [[f["name"], f.get("app", ""), f.get("space", ""), mb(f.get("byte_size")),
           f.get("cardinality")] for f in arr.get("largest_fields", [])]),
        ("App reload - 1-value fields",
         ["Field", "Single-value in # apps", "Spaces (sample)", "Sample app", "Rows", "Note"],
         [[f["field"], f["app_count"], f["spaces"], f["sample_app"], f["rows"],
           ("date field - likely the 1901-01-01 null placeholder; fix/drop at source"
            if f["date"] else "constant / dead column - candidate to drop (verify)")]
          for f in arr.get("single_value_fields", [])]),
        ("Import - largest datasets",
         ["Dataset", "Space", "Owner", "Size (MB)", "Rows", "Last load"],
         [[d["name"], d.get("space", ""), d.get("owner", ""), mb(d.get("size_bytes")),
           d.get("row_count"), str(d.get("last_load", ""))[:19]]
          for d in irr.get("largest_datasets", [])]),
        ("Import - largest files",
         ["File", "Space", "Owner", "Size (MB)", "Modified"],
         [[f["name"], f.get("space", ""), f.get("owner", ""), mb(f.get("size_bytes")),
           str(f.get("modified", ""))[:19]] for f in irr.get("largest_files", [])]),
        ("Import - duplicate datasets",
         ["Base name", "Dataset", "Space", "Size (MB)"],
         [[c["base_name"], g["name"], g.get("space", ""), mb(g.get("size_bytes"))]
          for c in irr.get("duplicate_datasets", []) for g in c["items"]]),
        ("Import - duplicate files",
         ["Base name", "File", "Space", "Size (MB)"],
         [[c["base_name"], g["name"], g.get("space", ""), mb(g.get("size_bytes"))]
          for c in irr.get("duplicate_files", []) for g in c["items"]]),
        ("Access errors by space",
         ["Space", "Apps errored", "Access denied (403/401)",
          "No metadata (404/never reloaded)", "Other", "Sample error"],
         [[e["space"], e["errored"], e["access_denied"], e["no_metadata"],
           e["other"], e["sample"]] for e in space_errors]),
    ]
    if orph:
        sheets += [
            ("Import - orphan files", ["File", "Size (MB)"],
             [[r["name"], mb(r.get("size_bytes"))] for r in orph.get("orphan_files", [])]),
            ("Import - produced unread", ["File", "Size (MB)"],
             [[r["name"], mb(r.get("size_bytes"))] for r in orph.get("produced_only_files", [])]),
            ("Import - orphan datasets", ["Dataset", "Space", "Size (MB)"],
             [[r["name"], r.get("space", ""), mb(r.get("size_bytes"))]
              for r in orph.get("orphan_datasets", [])]),
        ]
    if cons.get("periods"):
        sheets.append((
            "Consumption (raw)",
            ["Period type", "Start", "End", "Resource", "Action", "Task",
             "Capacity limit", "Local usage", "Overage", "Segments"],
            [[p.get("periodType", ""), str(p.get("periodStart", ""))[:10],
              str(p.get("periodEnd", ""))[:10], p.get("resourceType", ""),
              p.get("resourceAction", ""), p.get("taskName", ""), p.get("capacityLimit"),
              p.get("localUsage"), "Yes" if p.get("overage") else "",
              ", ".join(f"{k}={v}" for k, v in (p.get("segments") or {}).items())]
             for p in cons.get("periods", [])]))

    note = ("Capacity savings candidates - VERIFY before deleting. START with the 'Action list "
            "(billable)' sheet (archive / stale / duplicate apps ranked first). Only PERSONAL space "
            "is excluded (it does not count). Archive/shared/managed apps DO count - an app keeps "
            "billing its last-reload data size until it is reloaded smaller or DELETED. 'Counts?' "
            "shows whether the app loads EXTERNAL data (yes = counts; ~0 = only reads QVDs / binary "
            "loads; review = unclear). App 'Size' is the in-memory footprint, a PROXY for the billed "
            "Data-for-Analysis volume shown above.")

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except Exception:
        paths = []
        allsheets = [("summary", ["Metric", "Value"], summary_rows)] + \
                    [(re.sub(r"[^a-z0-9]+", "_", t.lower()).strip("_"), h, r) for t, h, r in sheets]
        for suffix, headers, rows in allsheets:
            p = os.path.join(out_dir, f"{base}_{suffix}.csv")
            with open(p, "w", newline="", encoding="utf-8-sig") as f:
                w = csv.writer(f)
                w.writerow(headers)
                w.writerows(rows)
            paths.append(p)
        log("openpyxl not installed - wrote CSV files instead of one workbook.")
        return paths[0]

    wb = Workbook()
    head_font = Font(bold=True, color="FFFFFF")
    head_fill = PatternFill("solid", fgColor="315C6D")
    warn_font = Font(bold=True, color="B00020")
    wrap = Alignment(wrap_text=True, vertical="top")
    wide = {"App", "Dataset", "File", "Field", "Base name", "Last reload", "Last load",
            "Segments", "Task", "Owner", "Reason", "Source", "Notes"}

    ws = wb.active
    ws.title = "Summary"
    ws.append(["CAPACITY OVERVIEW - read first"])
    ws["A1"].font = warn_font
    ws.append([note])
    ws.cell(row=2, column=1).font = warn_font
    ws.cell(row=2, column=1).alignment = wrap
    ws.append([])
    ws.append(["Metric", "Value"])
    hdr = ws.max_row
    for c in (1, 2):
        ws.cell(row=hdr, column=c).font = head_font
        ws.cell(row=hdr, column=c).fill = head_fill
    for r in summary_rows:
        rn = ws.max_row + 1
        ws.append(r)
        # same guard as add_sheet: a cell starting with = + - @ (e.g. the
        # "== PROXY TOTALS ==" dividers) must be stored as text, not a formula
        for ci, val in enumerate(r, 1):
            if isinstance(val, str) and val[:1] in ("=", "+", "-", "@"):
                ws.cell(row=rn, column=ci).data_type = "s"
    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 52

    def add_sheet(title, headers, rows):
        s = wb.create_sheet(title[:31])
        s.append(headers)
        for c in range(1, len(headers) + 1):
            s.cell(row=1, column=c).font = head_font
            s.cell(row=1, column=c).fill = head_fill
        for r in rows:
            rn = s.max_row + 1
            s.append(r)
            for ci, val in enumerate(r, 1):
                if isinstance(val, str) and val[:1] in ("=", "+", "-", "@"):
                    s.cell(row=rn, column=ci).data_type = "s"
        for i, h in enumerate(headers, 1):
            letter = get_column_letter(i)
            if h in wide:
                s.column_dimensions[letter].width = 42
            else:
                s.column_dimensions[letter].width = max(12, min(28, len(str(h)) + 4))
        if rows:
            s.freeze_panes = "A2"
            s.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{s.max_row}"

    for title, headers, rows in sheets:
        add_sheet(title, headers, rows)

    out_path = os.path.join(out_dir, f"{base}.xlsx")
    wb.save(out_path)
    log(f"Capacity report -> {os.path.basename(out_path)}")
    return out_path


def scan_meters(tenant, api_key, log=print):
    """FAST path — just the two authoritative meters (tenant-admin). A handful of
    REST calls, seconds. Run this FIRST to see WHICH capacity is over (and by how
    much) before launching the full inventory. Returns the summarized consumption
    or None if the key lacks consumption access."""
    try:
        cons = summarize_consumption(fetch_consumption(tenant, api_key, log=log))
    except Exception as e:
        log(f"Consumption endpoint unavailable (needs tenant-admin): {getattr(e, 'reason', e)}")
        return None
    dv = cons.get("data_volume")
    if dv:
        used, lim = dv.get("localUsage"), dv.get("capacityLimit")
        pct = f"{used / lim * 100:.3f}%" if (isinstance(used, (int, float)) and lim) else "?"
        over = (used - lim) if (isinstance(used, (int, float)) and isinstance(lim, (int, float))) else None
        flag = (f"  !! OVERAGE (over by {format_bytes(over)})"
                if dv.get("overage") and over and over > 0
                else ("  ! close to limit" if dv.get("closeToOverage") else ""))
        log(f"Data for Analysis (billed): {format_bytes(used)} / {format_bytes(lim)} ({pct}){flag}")
    else:
        log("Data for Analysis: no dataVolume record found (see the Consumption (raw) sheet).")
    # APP / QDI segment meters, only if this tenant emits them
    for lbl, rec in (("App reload (APP segment)", cons.get("app_reload")),
                     ("Import (QDI segment)", cons.get("import_qdi"))):
        if not rec:
            continue
        seg = rec.get("segments", {}) or {}
        val = next(iter(seg.values()), rec.get("localUsage"))
        lim = rec.get("capacityLimit")
        pct = f"{val / lim * 100:.3f}%" if (isinstance(val, (int, float)) and lim) else "?"
        log(f"{lbl}: {val} / {lim} ({pct})")
    return cons


if __name__ == "__main__":
    # Headless run. Needs a tenant-admin key for the authoritative consumption
    # numbers; degrades to proxies without it.
    #   QLIK_METERS_ONLY=1  -> just the two meters, in seconds (which one is over?)
    #   QLIK_ORPHANS=1      -> include orphan scan (only sensible on a small app set)
    TENANT = os.environ.get("QLIK_TENANT", "OUR.eu.qlikcloud.com")
    API_KEY = os.environ.get("QLIK_API_KEY", "")
    OUT_DIR = os.environ.get("QLIK_OUTPUT", ".")
    if not API_KEY:
        raise SystemExit("Set QLIK_API_KEY (and QLIK_TENANT) to run.")
    if os.environ.get("QLIK_METERS_ONLY"):
        scan_meters(TENANT, API_KEY, log=print)
    else:
        res = fetch_two_capacities(TENANT, API_KEY, log=print,
                                   with_orphans=bool(os.environ.get("QLIK_ORPHANS")))
        print_two_capacity_summary(res)
        write_capacity_report(res, OUT_DIR, log=print)
